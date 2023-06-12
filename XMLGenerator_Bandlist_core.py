# coding: utf-8

# IMPORTS
import openpyxl # openpyxl documentation: http://www.python-simple.com/python-autres-modules-non-standards/openpyxl.php
import string # ascii string constants library
from lxml import etree


# GLOBALS
xml_template = r"<?xml version='1.0' encoding='UTF-8'?><!-- Data type : BandlistSpecific notes :	General notes : 	- Most of the tags are optional, you can remove the really unnecessary ones.	- Tags marked as 'multiple' can be copied (with its block of sub-tag, up to the ending tag) if needed.	- all blocks marked 'OPTION' can be fully removed if not needed (now or in the future)	- **ABS MANDATORY / ABS COMPULSORY**: a value need to be absolutely provided, no way to escape! (SSHADE will not function properly if absent).	- **MANDATORY / COMPULSORY**: very important values for the search of the data. If the value (txt or numeric) of one tag is not known (or irrelevant in your case), then put 'NULL' and write a comment to keep track of the missing value. Remove comment when value is added.	- **MANDATORY / COMPULSORY only for ...**: when a value is optionally MANDATORY the condition is written. 	- 'LINK to existing UID' (unique identifier): references to another table in SSHADE. You have to reconstruct (easy for some: rule is in comment) or found this existing UID in the database beforehand (use 'Provider/Full Search' menu in SSHADE).	- 'UID to CREATE': you need to create this UID using their specific rules of creation that are explained in their attached comment. Use only alphanumeric characters and '_'.	- For UID you can use only alpha-numeric characters and the following: '_', '-'	- Enumeration type ('Enum' or 'OpenEnum') must contain one single item from the list given in brackets {}.	- use an optional <![CDATA[]]> tag when a value contains at least one special character (ie: &, >, <, /, ...). Example: <![CDATA[AT&T]]> for AT&T	- The data format is noted beetween [] and is 'ascii' when not specified. Ex: [Float], [Integer]. For [float] 2 formats are possible: decimal (123.456) or scientific (1.234e-56)  	- when no numerical format or Enum is specified, it is free text but limited to 256 characters. Only those noted [blob] have no size limitation.	- to import data for the first time you have to set <import_mode>='first import'. To correct data you have to change it to 'correction'.	- when a <filename> is given, then the file should be ziped with this xml file for import.--><import type='bandlist' ssdm_version='0.9.1'        xmlns='http://sshade.eu/schema/import'        xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'        xsi:schemaLocation='http://sshade.eu/schema/import http://sshade.eu/schema/import-0.9.xsd'>	<bandlist><!-- multiple -->	<!-- BAND LIST IMPORT MODE AND INDEXES -->		<import_mode>first import</import_mode> <!-- **ABS MANDATORY** Mode of import of 'bandlist' data. Enum: {first import, ignore, draft, no change, correction, invalidate} -->		<uid>BANDLIST_</uid> <!-- **ABS MANDATORY to CREATE** Unique identifier code given to the band list. Should be of the style ‘BANDLIST_Type_Molecule_Constituent’.  ‘Type’: 'ABS-R’, ‘REFL’, ‘THERM’ or ‘FLUO’; ‘Molecule’: general formula; ‘Constituent’: short description of the constituent (alphanumeric and '_', '-') -->	<!-- BAND LIST DESCRIPTION -->		<type></type> <!-- **ABS MANDATORY** Type of band list. Enum: {absorption, Raman scattering, reflectance, thermal emission, fluorescence emission} -->		<title></title> <!-- **ABS MANDATORY, requested for DOI** Title describing the band list. It should contain complete info on the band list type, the spectral range types, the specie, the constituent type and composition. It can also contain info on the environment parameters. This title is very important as it will be the title of the data reference generated from the DOI -->		<description><![CDATA[]]></description> <!-- **MANDATORY, recommended for DOI** Detailled description of the band list. It should contain complete info on the band list type, on typical spectral ranges covered, on the specie and constituent of this band list, as well as on the environment parameters (with values or range of values). Information on possible variations of the constituent may be also given [blob] -->	<!-- BANDLIST: SPECTRAL RANGES -->		<parameters_spectral> <!-- **ABS MANDATORY** -->			<unit></unit> <!-- **ABS MANDATORY** Unit of the position and width spectral parameters of the bands. Enum: {m-1, cm-1, angstrom, nm, micron, mm, m, km, Hz, kHz, MHz, GHz, eV, keV} -->			<standard></standard> <!-- **ABS MANDATORY** Medium in which the wavenumber/wavelength/frequency scale is given. Enum: {vacuum, air, unknown} -->			<range_types> <!-- **ABS MANDATORY at least one** -->				<type></type><!-- multiple --> <!-- **ABS MANDATORY** Typical spectral range (see equivalence table for typical minimum and maximum wavenumbers/wavelength/frequency). Enum: {gamma, hard X, soft X, EUV, VUV, UV, Vis, NIR, MIR, FIR, sub-mm, mm, cm, UHF, VHF, HF, MF, LF, VLF, ULF, SLF, ELF} -->			</range_types> 			<ranges> <!-- **ABS MANDATORY at least one** -->				<range><!-- multiple -->					<min></min> <!-- **ABS MANDATORY** Start of the spectral range of the band list. [Float] Unit: in “spectral_unit” -->					<max></max> <!-- **ABS MANDATORY** End of the spectral range of the band list. [Float] Unit: in “spectral_unit” -->				</range>			</ranges>												<comments><![CDATA[]]></comments> <!-- Additional information on spectral parameters: special range type, ... [blob] -->		</parameters_spectral>			<!-- BANDLIST: INTENSITY MODES AND UNITS -->		<reference_position>			<electronic></electronic> <!-- **MANDATORY for electronic bands when band_xxx_intensity_relative used** Reference position of the band to which the relative intensities of electronic bands are calculated (in 'spectral_unit') [Float] -->			<infrared></infrared> <!-- **MANDATORY for infrared and Raman bands when band_xxx_intensity_relative used** Reference position of the band to which the relative intensities of infrared active bands are calculated (in 'spectral_unit') [Float] -->		</reference_position>	<!-- BAND LIST: CONSTITUENT AND SPECIE -->		<constituent>			<uid></uid> <!-- **ABS MANDATORY** LINK to the existing UID of the “(basic) constituent” of the band list ['CONST_AB_yyyymmdd_123...']. The description of the constituent should reflect any range of composition values -->			<primary_specie_uid></primary_specie_uid> <!-- **ABS MANDATORY and only for molecular solids** LINK to the existing UID of the “primary molecular specie” of the constituent of the band list ['CONST_AB_yyyymmdd_123...'] -->			<comments><![CDATA[]]></comments> <!-- Additional information on the constituent (fundamental phase , …) and on the specie (composition, state, ...), in particular describe the parameters that change over the band list and the range of values spanned for the different bands [blob] -->		</constituent>			<!-- BANDLIST: HISTORY AND VERSION -->		<previous_version> 			<status></status> <!-- **ABS MANDATORY and only for band_import_mode={new version, invalidate}**  Validity status of the previous version of the band list. Enum: {obsolete version, partly invalidated version, invalidated version, partly invalidated data, invalidated data) -->			<comments></comments> <!-- **ABS MANDATORY and only for band_import_mode={new version, invalidate}** Description of the reason for the change of version or for the data invalidation [blob] -->		</previous_version>		<history><![CDATA[]]></history> <!-- Short description of the history of bandlist import, upgrade or correction (1 line max) that will be added to the 'date + import mode + version number' default info [blob] -->			<!-- BANDLIST: PARENT SPECTRA -->		<parent_experiments> <!-- **OPTION - STRONGLY RECOMMENDED when exist in the database** -->			<uid></uid><!-- multiple --> <!-- LINK to the existing UID of the original experiment (if in the database) used to produce part or all this band list ['EXPERIMENT_AB_yyyymmdd_123'] -->		</parent_experiments>		<parent_spectra> <!-- **OPTION - STRONGLY RECOMMENDED when exist in the database** -->			<uid></uid><!-- multiple --> <!-- LINK to the existing UID of the original spectra (if in the database) used to produce part or all this band list ['SPECTRUM_AB_yyyymmdd_123...'] -->		</parent_spectra>		<parent_spectra_comments><![CDATA[]]></parent_spectra_comments> <!-- Comments on parent spectra of the band list [blob] -->		<!-- BANDLIST: ANALYSIS AND VALIDATION -->		<analysis><![CDATA[]]></analysis> <!-- **MANDATORY** General description of the band analysis method(s) for the bandlist (ex: spectrum baseline correction + band fit (function, wavelength range, …); different sources of data, ...) [blob] -->		<comments><![CDATA[]]></comments> <!-- General comments on the bandlist, measurements conditions or analysis [blob] -->		<quality_flag></quality_flag> <!-- **MANDATORY** Global quality flag on the band list. Can have mostly 6 quality levels from ‘0’: “no valid data” to ‘5’: “excellent data”. Enum: {0, 1, 2, 3, 4, 5} [Integer] -->		<date_validated></date_validated> <!-- **MANDATORY** Validation date of the version of the band list [Format: 'YYYY-MM-DD'] -->		<validators> <!-- **MANDATORY** -->			<experimentalist_uid></experimentalist_uid><!-- multiple --> <!-- **MANDATORY** LINK to the existing UID of the experimentalist(s) who processed and validated this version of the band list  [‘EXPER_Firstname_Lastname(_n)'] -->		</validators>	<!-- BANDLIST: DOCUMENTATIONS -->		<documentations> <!-- **OPTION** Documentations about the band list -->			<documentation><!-- multiple -->				<name><![CDATA[]]></name> <!-- Name of the documentation -->				<filename><![CDATA[]]></filename> <!-- File name (.pdf, ...) of documentation. This file should be zipped with the xml file -->			</documentation>		</documentations>	<!-- BANDLIST: FILES -->		<original_data_filename></original_data_filename> <!-- Name of the file (with extension: .txt, .doc, .rtf, ...) containing the original data of the band list. This file should be zipped with the xml file -->		<export_filename></export_filename> <!-- **ABS MANDATORY** Generic name of the band list data and metadata files and preview image for data export (will get different extensions: .dat, .txt, .xml… depending on the export format chosen by user). Do not use blank(space) in this name -->			<!-- BANDLIST PREVIEW -->		<preview>			<x axis='' unit='' min='' max=''/> <!-- Type of X axis and min/max values in the preview plot of the band list. For 'axis': Enum: {lin, log}; For 'unit': Enum: {m-1, cm-1, angstrom, nm, micron, mm, m, km, Hz, kHz, MHz, GHz, eV, keV}; For 'min/max': [float] -->			<y axis='' unit='' min='' max=''/> <!-- Type of Y axis, unit, and min/max values in the absorption coefficient preview plot of the band list. For 'axis': Enum: {lin, log}; For 'unit': OpenEnum: {cm-1}; For 'min/max': [float] -->			<yrel axis='' min='' max=''/> <!-- Type of Y axis and min/max values in the relative intensity preview plot of the band list. For 'axis': Enum: {lin, log}; For 'min/max': [float] -->			<type></type> <!-- Flag telling which plot is used as the preview, in place of the default one. Enum: {absorption coefficient, relative intensity} -->		<!-- OR -->			<filename></filename> <!-- **Include the full 'bandlist_title' on top of the figure and provide the image file with a size around 650x460 pixels** Name of the file (with only .png, .jpg, (.gif) extensions) containing the preview plot of the band list to be displayed (otherwise automatically generated during import). This image file should be zipped with the xml file -->		</preview>	<!-- BANDLIST: STRUCTURE -->			<structure> <!-- **STRONGLY RECOMMENDED when there are several types of transitions as well as isotopic species** -->			<!-- This structure is used to organize the bandlist in sections and subsections, in particular when we want to separate ‘absorption’ bands and ‘Raman’ bands, and also to separate the different isotopes of a molecule. The bandlist, the sections and the subsections will be displayed in SSHADE in the order and with the structure defined here together with their title and the information on their type of variable parameters (band type, isotope, band transition category). The section and subsection titles are subtitles of the general title of the bandlist ('bandlist_title') -->		<!-- SECTIONS -->			<sections variable_parameter=''> <!-- **MANDATORY** Type of bandlist, band or constituent parameter varying between the sections of the band list. OpenEnum: {no, isotope, band transition category, other} -->				<section><!-- multiple: at least 2 -->					<title></title> <!-- **MANDATORY** Title describing the band type, isotope, or band transition category of this section of the band list -->					<description><![CDATA[]]></description> <!-- Detailled description of the band type, isotope, or band transition category of this section of the band list [blob] -->				<!-- BANDS --> <!-- **OPTION #1** -->					<!-- Use option when there are only two varying parameters in the band list -->					<bands>						<band_uid></band_uid><!-- multiple --> <!-- **ABS MANDATORY at least one** Link to the existing UID of the band belonging to this band list section. List them in decreasing order of band position ['BAND_Type_Molecule_Constituent_Wavenumber'] -->					</bands>									<!-- or -->										<!-- SUBSECTIONS --> <!-- **OPTION #2** -->					<subsections variable_parameter=''> <!-- **MANDATORY for option #2** Type of bandlist, band or constituent parameter between the subsections of this section of the band list. OpenEnum: {no, isotope, band transition category, other} -->						<subsection><!-- multiple: at least 2 -->							<title></title> <!-- **MANDATORY** Title describing the band type, isotope, or band transition category of this subsection of the band list -->							<description><![CDATA[]]></description> <!-- Detailled description of the band type, isotope, or band transition category of this subsection of the band list [blob] -->						<!-- BANDS -->							<bands>								<band_uid></band_uid><!-- multiple --> <!-- **ABS MANDATORY at least one** Link to the existing UID of the band belonging to this band list subsection. List them in decreasing order of band position ['BAND_Type_Molecule_Constituent_Wavenumber'] -->							</bands>						</subsection>					</subsections>									</section>			</sections>		</structure><!-- ******************************************************************* -->	<!-- BANDS -->		<bands>			<band><!-- multiple -->			<!-- BAND: IMPORT MODE AND INDEXES -->				<import_mode>first import</import_mode> <!-- **ABS MANDATORY** Mode of import of 'band' data. Enum: {first import, inherited, ignore, draft, no change, correction, new version, invalidate} -->				<uid>BAND_</uid> <!-- **ABS MANDATORY to CREATE** Unique identifier code given to the band. Should be of the style ‘BAND_Type_(Molecule_)Constituent_Wavenumber(_a)’.  ‘Type’: 'ABS’, 'RAMAN', ‘REFL’, ‘THERM’ or ‘FLUO’; ‘Molecule’: formula (not isotopic); ‘Constituent’: short description of the constituent; ‘Wavenumber’ approximate wavenumber (cm-1) or wavelength (nm: then add 'nm'); ‘a’: optional extension [alphanumeric and '_', '-'] -->			<!-- BAND: COMMENTS -->				<comments><![CDATA[]]></comments> <!-- General comments on the band, measurements conditions (type of spectrum) or analysis [blob] -->			<!-- BAND: TRANSITION ASSIGNMENTS AND MODES -->				<assignments>					<assignment>						<number></number> <!-- **ABS MANDATORY only when more than one** Order number of the assignment of the band [integer] -->						<label></label> <!-- **ABS MANDATORY** Symbolic label of the band transition [LaTEX format]. Should also include the symmetry (character) of the vibration for phonon and libration modes. Ex: '$2\nu_1 + 3\nu_2$'; '$R_{xy} (T_g)$'; ('?' or '?+?' for unknown) -->						<symmetry_label></symmetry_label> <!-- **MANDATORY for fundamental vibration, rotation & phonon mode}** Symmetry (character) of the vibration producing the band in the molecular solid structure. OpenEnum: {A, Ag, Au, Ap, As, A1, A1g, A1u, A1p, A2, A1s, A2g, A2u, A2p, A2s, B, Bg, Bu, B1, B1g, B1u, B2, B2g, B2u, B3, B3g, B3u, E, Eg, Eu, Ep, Es, E1, E1g, E1u, E2, E2g, E2u, F, Fg, Fu, F1, F1g, F1u, F2, F2g, F2u, T, Tg, Tu, T1, T1g, T1u, T2, T2g, T2u, other, unknown} -->						<category></category> <!-- **ABS MANDATORY** General category of transition producing this band. Enum: {electronic transition, fundamental vibration, overtone vibration, combination vibration, two-phonon mode, vibron-phonon mode, rotation, overtone rotation, combination, phonon mode, other, unknown} -->						<method></method> <!-- Description of the method of band transition assignment [blob] -->					<!-- TRANSITION ASSIGNMENT: EVALUATION -->						<level></level> <!-- **MANDATORY** Level of assignment of the specie(s) and transition(s) contributing to the band. Enum: {fully assigned, partly assigned, transition assigned, species assigned, uncertain assignment, not assigned} -->						<evaluation></evaluation> <!-- **MANDATORY** Evaluation of the band assignment. Enum: {undefined, uncertain, validated, recommended, with caution, not recommended} -->						<comments><![CDATA[]]></comments> <!-- Comments on the band transition assignment [blob] -->					<!-- TRANSITION: MULTIPLICITY AND DEGENERACY -->						<multiplicities> 							<multiplicity><!-- multiple -->								<type></type> <!-- **ABS MANDATORY** Type of multiple contributions to this band. Enum: {no, mode degeneracy, site degeneracy, rotational structure, accidental degeneracy, other isotope specie, other constituent specie, other, unknown} -->								<degeneracy></degeneracy> <!-- **MANDATORY for mode/site/accidental degeneracy** Degeneracy (mode, site or accidental) of the vibration mode of the isotope species producing this band. Enum: {no, double, triple, quadruple, double site, triple site, accidental double, accidental triple, other, unknown} -->								<other_band_uid></other_band_uid> <!-- **MANDATORY when 'contribution_level' is 'extracted'** Link to the existing UID of the other band contributing at the same position as this band -->							</multiplicity>						</multiplicities> 						<contribution_level></contribution_level> <!-- **MANDATORY except for multiplicity_type = no/mode/site degeneracy** Qualitative level of contribution of this transition to the band. Enum: {full, extracted, major, medium, minor, unknown}. Default = 'full' -->						<contribution_comments><![CDATA[]]></contribution_comments> <!-- Comments on the contribution of this transition to the band characteristics [blob] -->					<!-- BAND: TRANSITION MODES -->						<transition> <!-- **ABS MANDATORY at least one type of transition mode** -->						<!-- The transition is multiple only in case of combinations of modes. Each mode of the combination will be described -->							<primary_species>								<primary_specie>									<uid></uid><!-- multiple --> <!-- **ABS MANDATORY except for phonon modes** Link to the existing UID of the “primary specie” subjected to the transition that produces the band ['ATOM_', 'ATION_', ‘MOLEC_(Isotopic)Chemicalformula(Letter)’, ‘MOLION_(Isotopic)ChemicalformulaCharge(Letter)’] -->									<crystal_molecule_sites> <!-- **OPTION, ONLY for molecular solids** Crystallographic site of the molecular species producing the band in the molecular solid structure -->										<site><!-- multiple --> 											<label></label> <!-- **COMPULSORY for OPTION** Label of the crystallographic site of the molecular species. This site should be already described in the constituent. FreeList: {A, B, C…, I, I1, I2…, cage 5-12, cage 5-12_6-4, cage 5-12_6-4, cage 4-3_5-6_6-3, cage 5-12_6-8, …} -->											<symmetry_label></symmetry_label> <!-- **COMPULSORY for OPTION** Symmetry of the molecular species producing the band in the molecular solid structure. FreeList: {A1, A2, B1, B2, …} -->										</site>									</crystal_molecule_sites>									<crystal_sites> <!-- **OPTION, ONLY for ionic/covalent solids** Crystallographic site of the atomic species producing the band in the ionic/covalent constituent structure -->										<label></label><!-- multiple --> <!-- **COMPULSORY for OPTION** Label of the crystallographic site of the atomic species. This site should be already described in the constituent. FreeList: {M1, M2, M3, M4, O1, O2, O3, O4, O5, O6, O7, O8, …} -->									</crystal_sites>									<sites_comments><![CDATA[]]></sites_comments> <!-- Comments on the sites of this specie [blob] -->								</primary_specie>							</primary_species>						<!-- TRANSITIONS -->							<!-- **ABS MANDATORY at least one mode block option** **ONLY ONE 'MODE BLOCK' PER TRANSITION TYPE -->						<!-- TRANSITION: ELECTRONIC -->							<electronic_modes> <!-- OPTION #1 - MANDATORY for electronic transitions -->								<mode><!-- multiple -->										<label></label> <!-- **MANDATORY** Label including the lower and upper states of the electronic transition of the specie. Ex: $^6A_1 ~ \rightarrow ~ ^4E_1$. Format: LaTeX -->									<type></type> <!-- **MANDATORY** Type of electronic transition mode of the specie(s) producing this band. Enum: {atomic electronic transition, molecular electronic transition, crystal field, ligand-to-metal charge-transfer, intervalence charge transfer, double exciton, other, unknown} -->								</mode>								<comments><![CDATA[]]></comments> <!-- Comments on the electronic transition modes of the specie [blob] -->							</electronic_modes>						<!-- TRANSITION: VIBRATION MODES -->							<vibration_modes> <!-- OPTION #2 - MANDATORY for vibration transitions -->								<mode><!-- multiple -->										<label></label> <!-- **MANDATORY** Label of the normal mode of vibration of the molecular specie. ex: $2nu_2$, $2\nu_1 + \nu_3$. Format: LaTeX -->									<type></type> <!-- **MANDATORY** Type of vibration mode of the bond or molecule producing this band. Enum: {stretching, stretching sym., stretching asym., bending, bending in-p, bending out-p, bending sym., bending asym., bending sym. in-p (scissoring), bending asym. in-p (rocking), bending sym. out-p (wagging), bending asym. out-p (twisting), deformation, deformation in-p, deformation out-p, deformation sym., deformation asym., other, unknown}. Note: old [xxx antisym.]: replaced by 'xxx asym.' -->									<chemical_bonds>										<uid></uid><!-- multiple --> <!-- **MANDATORY** Link to the existing UID of the bond, part of the molecule, or whole molecule, of the primary species or ionic/covalent constituent subjected to this mode of vibration. [‘BOND_ZAtom1(BondSymbol)ZAtom2(Charge)’ or ‘MOLECPART_’, ‘MOLEC_(Isotopic)Chemicalformula(Letter)’, ‘MOLION_(Isotopic)ChemicalformulaCharge(Letter)’, 'MOLRAD_' or 'MOLRADION_'] -->									</chemical_bonds>								</mode>								<comments><![CDATA[]]></comments> <!-- Comments on the vibration modes of the specie [blob] -->							</vibration_modes>						<!-- TRANSITION: ROTATION MODES -->							<rotation_modes> <!-- OPTION #3 - **MANDATORY for rotation transitions** -->								<mode><!-- multiple -->										<label></label> <!-- **MANDATORY** Label of the normal mode of rotation of the molecular specie. FreeList: {$\nu_R$, $\nu_{Rx}$, $\nu_{Ry}$, $\nu_{Rxy}$, $\nu_{Rz}$, $\nu_{Rxyz}$, $\nu_L$, $\nu_{Lx}$, $\nu_{Ly}$, $\nu_{Lxy}$, $\nu_{Ly}$, $\nu_{Lxyz}$, ?}. Format: LaTeX -->									<type></type> <!-- **MANDATORY** Type of rotation mode of the molecule producing this band. Enum: {free rotation, hindered rotation, libration, other, unknown} -->								</mode>								<comments><![CDATA[]]></comments> <!-- Comments on the rotation modes of the specie [blob] -->							</rotation_modes>						<!-- TRANSITION: PHONON MODES -->							<phonon_modes> <!-- OPTION #4 - **MANDATORY for phonon transitions** -->								<mode><!-- multiple -->									<label></label> <!-- **MANDATORY** Label of the translation phonon mode of the solid contributing to this band. FreeList: {$nu_T$, $nu_{Txy}$, $nu_{Tz}$, $nu_{T(LO)}$, $nu_{T(LOx)}$, $nu_{T(LOy)}$, $nu_{T(LOy)}$,$nu_{T(TO)}$, $nu_{T(LA)}$, $nu_{T(TA1)}$, $nu_{T(TA2)}$, ?}. Format: LaTeX -->									<type></type> <!-- **MANDATORY** Type of phonon mode of the solid structure producing this band. Enum: {translation, longitudinal optic translation, transverse optic translation, longitudinal acoustic translation, transverse acoustic translation, other, unknown} -->								</mode>								<comments><![CDATA[]]></comments> <!-- Comments on the phonon modes of the specie [blob] -->							</phonon_modes>						</transition>					<!-- TRANSITION: RESONANCES -->						<resonances> <!-- **COMPULSORY when exist** Transition modes (internal or external) with which this mode is in resonance -->							<resonance><!-- multiple -->								<type></type> <!-- **MANDATORY** Type of resonance of the transition mode. Enum: {Fermi resonance, electron-phonon coupling, rotational-vibrational coupling, vibration-phonon coupling, other, unknown} -->								<band_uid></band_uid> <!-- **MANDATORY** Link to the existing UID of the band which contain the transition mode in resonance with this transition. [‘BAND_Type_Molecule_Constituent_Wavenumber'] -->								<band_assignment_number></band_assignment_number> <!-- **MANDATORY** Order number of the assignment of the transition (of the band above) which is in resonance [integer] -->								<comments><![CDATA[]]></comments> <!-- Comments on the resonance of this transition [blob] -->							</resonance>						</resonances>					</assignment>				</assignments>			<!-- BAND: ENVIRONMENT AND CHARACTERISTICS -->				<characteristics>					<characteristic>						<number></number> <!-- **ABS MANDATORY only when more than one** Order number of the characteristic of the band [integer] -->					<!-- BAND: CONSTITUENT -->						<constituent> <!-- Describe the effective composition, phase and texture of the constituent for this set of band characteristics -->							<composition_comments><![CDATA[]]></composition_comments> <!-- **COMPULSORY when different from original constituent** Effective values of the constituent composition (and phase), for this band when changed (or precised) compared to the nominal constituent [blob] -->							<texture_comments><![CDATA[]]></texture_comments> <!-- **COMPULSORY when different from original constituent** Constituent texture for this band when it affect its band characteristics (mostly reflectance & emission band lists) [blob] -->													<!-- BAND: CONSTITUENT ENVIRONMENT PARAMETERS -->							<parameters_environment>							<!-- Describe the environment parameters of the constituent for this set of band characteristics -->							<!-- CONSTITUENT ENVIRONMENT: TEMPERATURE -->								<temperature>									<unit></unit> <!-- **ABS MANDATORY** Unit of temperature. Enum: {K, C, F} -->									<value></value> <!-- **ABS MANDATORY** Effective temperature of the constituent. [Float] Unit: in “temperature_unit” -->									<error></error> <!-- **MANDATORY** Absolute uncertainty on effective temperature of the constituent. [Float] Unit: in “temperature_unit” -->									<formation></formation> <!-- Formation temperature of the constituent. [Float] Unit: in “temperature_unit” -->									<max></max> <!-- **MANDATORY** Maximum temperature (annealing, ...) reached by the constituent after formation. [Float] Unit: in “temperature_unit” -->									<comments><![CDATA[]]></comments> <!-- Comments about the temperature environment and history of the constituent [blob] -->								</temperature>							  							<!-- CONSTITUENT ENVIRONMENT: MECHANICAL PRESSURE APPLIED TO THE CONSTITUENT -->								<pressure> <!-- **OPTION** -->									<unit></unit> <!-- **ABS MANDATORY when <pressure_value>, <formation> or <max> has a value** Unit of mechanical pressure. Enum: {Pa, hPa, MPa, GPa, mbar, bar, atm, torr} -->									<value></value> <!-- **MANDATORY for OPTION** Effective mechanical pressure applied to the constituent. [Float] Unit: in “pressure_unit” -->									<error></error> <!-- Absolute uncertainty on the effective mechanical pressure. [Float] Unit: in “pressure_unit” -->									<formation></formation> <!-- Mechanical pressure applied to the constituent during formation. [Float] Unit: in “pressure_unit” -->									<max></max> <!-- Maximum mechanical pressure applied to the constituent after formation. [Float] Unit: in “pressure_unit” -->									<stress_type></stress_type> <!-- Type of stress applied on the constituent after formation.  Enum: {normal uniaxial tension, normal uniaxial compression, simple shear, normal biaxial tension, normal biaxial compression, cylindrical normal tension, cylindrical normal compression, isotropic normal tension, isotropic normal compression, combined biaxial, combined triaxial, other, unknown} -->									<comments><![CDATA[]]></comments> <!-- Comments about the pressure and stress environment and history of the constituent [blob] -->								</pressure>							</parameters_environment>						</constituent>					<!-- BAND: EXCITATION LIGHT -->						<excitation> <!-- Describe the excitation light for this set of band characteristics -->							<laser_wavelength><![CDATA[]]></laser_wavelength> <!-- **MANDATORY for Raman and fluorescence** Wavelength of the laser excitation [Float] Unit: in 'nm' -->							<sample_orientation_mode></sample_orientation_mode> <!-- **MANDATORY for Raman** Orientation mode of the axe of the incident excitation light relative to the crystalline axes of the constituent. Enum: {oriented, unoriented, random, unknown} -->							<sample_orientation></sample_orientation> <!-- **MANDATORY for Raman with oriented sample** Orientation of the axe of the incident excitation light relative to the crystalline axes of the constituent -->							<polarization_orientation_mode></polarization_orientation_mode> <!-- **MANDATORY for Raman and absorption with oriented sample** Orientation mode of the polarization of the incident excitation light relative to the crystalline axes of the constituent. Enum: {depolarized, polarized, unknown} -->							<polarization_orientation></polarization_orientation> <!-- **MANDATORY for polarized light** Orientation of the polarization of the incident excitation light relative to the crystalline axes of the constituent -->							<comments><![CDATA[]]></comments> <!-- Comments about the pressure and stress environment and history of the constituent [blob] -->						</excitation>					<!-- BAND: CHARACTERISTICS METHODS -->						<methods> <!-- **MANDATORY at least one** General methods used to get the values of the band characteristics -->							<method><!-- multiple -->								<type></type> <!-- **MANDATORY** Type of method used to get the values of the band characteristics from the spectrum. Enum: {spectrum measurement, spectrum fit, spectrum analysis, data compilation, data extrapolation, theory, estimation, various, other, unknown} -->								<description></description> <!-- Description of the method used to get the values of the different band characteristics [blob] -->							<!-- BAND: FIT FUNCTION -->								<fit_function> <!-- **OPTION for 'spectrum fit'** -->									<type></type> <!-- Type of shape function used to fit the band. OpenEnum: {Gaussian, Voigt, Lorentzian, BWF, Doppler, other, unknown} -->									<parameters><![CDATA[]]></parameters> <!-- List of the parameters (name and symbol) of the fit function and their values (with unit) for the band. Ex: 'central frequency $\omega_0 = 1254 cm^{-1}$, FWHM $\Delta_G = 23 cm^{-1}$, intensity $I_0 = 42500 cm^{-1}$' [blob] -->								</fit_function>							</method>						</methods>					<!-- BAND OVERLAP -->						<overlap></overlap> <!-- **MANDATORY** Degree of overlapping of the band with another band. Enum: {extracted, isolated, slightly blended, moderately blended, strongly blended, shoulder-tail, multiple, other, unknown} -->					<!-- BAND POSITION -->						<position>							<peak_method></peak_method> <!-- **ABS MANDATORY when value in 'peak'** Method of determination of the band peak position. Enum: {peak, fit peak, 90%-max center, first derivative, second derivative, higher order derivative, extrapolated, calculated, estimated, various, other, unknown} -->							<peak></peak> <!-- **ABS MANDATORY or 'center'** Position of the peak or center of the band (depending on method). [Float] Unit: in “bandlist_spectral_unit” -->							<peak_error></peak_error> <!-- **ABS MANDATORY when value in 'peak'** Absolute uncertainty (+/-) in the position of the band peak or center. [Float] Unit: in “bandlist_spectral_unit” -->							<center_method></center_method> <!-- **ABS MANDATORY when value in 'center'** Method of determination of the band center position. Enum: {half-max center, fit center, first derivative, second derivative, higher order derivative, extrapolated, calculated, estimated, various, other, unknown} -->							<center></center> <!-- **ABS MANDATORY or 'peak'** Position of the band center at half maximum intensity. [Float] Unit: in “bandlist_spectral_unit” -->							<center_error></center_error> <!-- **ABS MANDATORY when value in 'center'** Absolute uncertainty (+/-) of the position of the band center at half maximum intensity. [Float] Unit: in “bandlist_spectral_unit” -->							<!-- <spectral_range_type></spectral_range_type> --> <!-- **ABS MANDATORY - Calculated from 'peak'** Typical spectral range of the band. Enum: {gamma, hard X, soft X, EUV, VUV, UV, Vis, NIR, MIR, FIR, sub-mm, mm, cm, UHF, VHF, HF, MF, LF, VLF, ULF, SLF, ELF} -->							<evaluation></evaluation> <!-- **ABS MANDATORY** Evaluation of the value of band position. Enum: {undefined, uncertain, validated, recommended, with caution, not recommended} -->							<!-- <quality_flag></quality_flag> --> <!-- **CALCULATED** Quality flag on band position accuracy. Enum: {A, B, C, D, E, undefined}. Accuracy: A: 0~1, B: 1~3, C: 3~6, D: 6-15, E: 15-30 cm-1 -->							<comments><![CDATA[]]></comments> <!-- Comments on the band position and accuracy [blob] -->						</position>												<!-- BAND WIDTH -->						<width>							<method></method> <!-- **MANDATORY when value in 'fwhm'** Method of determination of band width. Enum: {fwhm, fit fwhm, hwhm, first derivative, extrapolated, calculated, estimated, various, other, unknown} -->							<fwhm></fwhm> <!-- **MANDATORY** Full width at half maximum (FWHM) of the band. [Float] Unit: in “bandlist_spectral_unit” -->							<fwhm_error></fwhm_error> <!-- **MANDATORY when value in 'fwhm'** Absolute uncertainty (+/-) of the full width at half maximum of the band. [Float] Unit: in “bandlist_spectral_unit” -->							<shape></shape> <!-- **MANDATORY** Type of band shape. OpenEnum: {symmetric, gaussian, lorentzian, Breit-Wigner-Fano, voigt, doppler, asymmetric, asymmetric low frequency wing, asymmetric high frequency wing, shoulder, sharp shoulder, broad shoulder, low frequency tail, high frequency tail, undefined, other, unknown} -->							<asymmetry_factor></asymmetry_factor> <!-- **MANDATORY for asymmetric band** Asymmetry factor of the band shape [Float] -->							<asymmetry_factor_error></asymmetry_factor_error> <!-- Absolute uncertainty (+/-) of the asymmetry factor of the band shape [Float] -->							<evaluation></evaluation> <!-- **MANDATORY when value in 'fwhm'** Evaluation of the value of band width. Enum: {undefined, uncertain, validated, recommended, with caution, not recommended} -->							<!-- <quality_flag></quality_flag> --> <!-- **CALCULATED when value in 'fwhm_error'** Quality flag on band width accuracy. Enum: {A, B, C, D, E, undefined}. Accuracy: A: 0~0.3, B: 0.3~1, C: 1~2, D: 2-5, E: 5-10 cm-1 -->							<comments><![CDATA[]]></comments> <!-- Comments on the band width [blob] -->						</width>										<!-- BAND INTENSITY -->						<peak_intensity>							<method></method> <!-- **MANDATORY when block used** Method of determination of band intensity. Enum: {peak intensity, baseline corrected peak intensity, fit intensity, extrapolated, calculated, estimated, various, other, unknown} -->							<abscoef></abscoef> <!-- **MANDATORY only for 'absorption', or fill 'relative'** Absorption coefficient value at band peak. [Float] Unit: in 'cm-1' -->							<abscoef_error></abscoef_error> <!-- **MANDATORY when value in 'abscoef' ** Absolute uncertainty (+/-) of the absorption coefficient intensity value at peak. [Float] Unit: in 'cm-1' -->							<abscoef_specific></abscoef_specific> <!-- Specific absorption coefficient of the species at peak (i.e. abscoef divided by its abundance). [Float] Unit: in 'cm-1' -->							<abscoef_specific_error></abscoef_specific_error> <!-- Absolute uncertainty (+/-) in specific absorption coefficient at peak. [Float] Unit: in 'cm-1' -->							<relative></relative> <!-- **MANDATORY, or fill 'abscoef'** Relative intensity value at band peak, compared to the reference band (“bandlist_position_reference”). [Float] Provide the ratio, not in % -->							<relative_error></relative_error> <!-- **MANDATORY when value in 'relative'** Absolute uncertainty (+/-) of the relative peak intensity value. [Float] -->							<strength></strength> <!-- **MANDATORY** Qualitative (relative) spectroscopic classification of band peak intensity. Enum: {ia, ew, vvw, vw, w, m, s, vs, vvs, es, unknown}. For abscoef (relative) [reflectance]: ia: <0.0001 (<1e-8) [<0.0002], ew: 0.0001~0.01 (1e-8~1e-6) [0.0002~0.002], vvw: 0.01~0.1 (1e-6~1e-5) [0.002~0.005], vw: 0.1~1 (1e-5~1e-4) [0.005~0.02], w: 1~10 (1e-4~1e-3) [0.02~0.05], m: 10~100 (0.001~0.01) [0.05~0.25], s: 100~1000 (0.01~0.1) [0.25~0.75], vs: 1000~10000 (0.1~1) [0.75~1], vvs: 1e4-1e5 cm-1 (1-10) [1-1.25], es: >1e5 cm-1 (>10) [>1.25] -->							<evaluation></evaluation> <!-- **MANDATORY when value in 'abscoef' or in 'relative'** Evaluation of the value of band peak intensity. Enum: {undefined, uncertain, validated, recommended, with caution, not recommended} -->							<!-- <quality_flag></quality_flag> --> <!-- **CALCULATED when value in 'xxx_error'** Quality flag on accuracy of the band peak intensity. Enum: {A, B, C, D, E, undefined}. Accuracy: A: <2%, B: 2~5%, C: 5~10%, D: 10-25%, E: >25% -->							<comments><![CDATA[]]></comments> <!-- Comments on the band peak intensity [blob] -->						</peak_intensity>					<!-- BAND INTEGRATED INTENSITY -->						<integrated_intensity>							<method></method> <!-- **MANDATORY when block used** Method of determination of band integrated intensity. Enum: {band integrated intensity, width x peak intensity, fit integrated intensity, extrapolated, calculated, estimated, various, other, unknown} -->							<abscoef></abscoef> <!-- **MANDATORY only for 'absorption', or fill 'relative'** Absorption coefficient integrated over the band. [Float] Unit: in 'cm-2' -->							<abscoef_error></abscoef_error> <!-- **MANDATORY when value in 'abscoef'** Absolute uncertainty (+/-) in band intensity integrated over the band. [Float] Unit: in 'cm-2' -->							<abscoef_specific></abscoef_specific> <!-- Specific absorption coefficient of the species integrated over the band (i.e. abscoef divided by its abundance).. [Float] Unit: in 'cm-2' -->							<abscoef_specific_error></abscoef_specific_error> <!-- Absolute uncertainty (+/-) in specific absorption coefficient integrated over the band. [Float] Unit: in 'cm-2' -->							<relative></relative> <!-- **MANDATORY, or fill 'abscoef'** Relative integrated band intensity, compared to the reference band (“bandlist_position_reference”). [Float] Provide the ratio, not in % -->							<relative_error></relative_error> <!-- **MANDATORY when value in 'relative'** Absolute uncertainty (+/-) of the relative integrated band intensity. [Float] -->							<strength></strength> <!-- Qualitative (relative) spectroscopic classification of integrated band intensity. Enum: {ia, ew, vvw, vw, w, m, s, vs, vvs, es, unknown}. For abscoef (relative) [reflectance]: ia: <0.001 (<1e-8) [<0.0002], ew: 0.001-0.1 ((1e-8~1e-6) [0.0002~0.002], vvw: 0.1~1 (1e-6~1e-5) [0.002~0.005], vw: 1~10 (1e-5~1e-4) [0.005~0.02], w: 10~100 (1e-4~1e-3) [0.02~0.05], m: 100~1000 (0.001~0.01) [0.05~0.25], s: 1e3~1e4 (0.01~0.1) [0.25~0.75], vs: 1e4~1e5 (0.1~1) [0.75~1], vvs: 1e5-1e6 cm-2 (1-10) [1-1.25], es: >1e6 cm-2 (>10) [>1.25] -->							<evaluation></evaluation> <!-- **MANDATORY when value in 'abscoef' or in 'relative'** Evaluation of the value of band integrated intensity. Enum: {undefined, uncertain, validated, recommended, with caution, not recommended} -->							<!-- <quality_flag></quality_flag> --> <!-- **CALCULATED when value in 'xxx_error'** Quality flag on accuracy of the band integrated intensity. Enum: {A, B, C, D, E, undefined}. Accuracy: A: <2%, B: 2~5%, C: 5~10%, D: 10-25%, E: >25% -->							<comments><![CDATA[]]></comments> <!-- Comments on the band integrated intensity [blob] -->						</integrated_intensity>					<!-- BAND: BANDLIST PREVIEW -->						<bandlist_nominal_flag></bandlist_nominal_flag> <!-- **ABS MANDATORY** Flag defining if this set of band characteristics is part of the nominal band list - BoolEnum: {yes, no} or {true, false} -->					</characteristic>				</characteristics>			<!-- BAND: REFERENCES -->				<publications>					<publication_uid></publication_uid><!-- multiple --> <!-- **MANDATORY, or spectrum_uid, or link** LINK to the existing UID of the publication in which information on this band is published. [‘PUBLI_FirstAuthorName_Year(Letter)’] -->				</publications>				<data_publication_spectra>					<spectrum_uid></spectrum_uid><!-- multiple --> <!-- **MANDATORY, or publication or link** Link to the existing UID of the experiment from which information on this band has been extracted ['EXPERIMENT_AB_yyyymmdd_123'] -->				</data_publication_spectra>				<data_publication_links>					<link></link><!-- multiple --> <!-- **MANDATORY, or publication or spectrum_uid** Link to the existing DOI (or URL) of the spectrum or experiment from which information on this band has been extracted -->				</data_publication_links>				<publication_comments><![CDATA[]]></publication_comments> <!-- Comments about the band in the publication: which characteristics of the band is published, and in which publication ... [blob] -->			</band>		</bands>	</bandlist></import>"
letters = tuple(string.ascii_uppercase)
col_n = dict() # column dict for BandList sheet (only BandList, no Band)
col_n = {
    "B_Index": "A",
    "B_Import_mode": "B",
    "B_UID": "C",
    "B_Type": "D",
    "B_Comment": "E",
    "B_Assignment_Number": "H",
    "B_Assignment_Label": "I",
    "B_Assignment_Symmetry": "J",
    "B_Assignment_Category": "K",
    "B_Assignment_Method": "L",
    "B_Assignment_Level": "M",
    "B_Assignment_Evaluation": "N",
    "B_Assignment_Comment": "O",
    "B_Assignment_Multiplicity_Types": "R",
    "B_Assignment_Multiplicity_Degeneracy": "S",
    "B_Assignment_Multiplicity_Other_band": "T",
    "B_Assignment_Contribution_Level": "U",
    "B_Assignment_Contribution_Comment": "V",
    "B_Assignment_Transition_Specie_UID": "Y",
    "B_Assignment_Sites_Molecule_labels": "Z",
    "B_Assignment_Sites_Molecule_Symm_label": "AA",
    "B_Assignment_Sites_Atom_Labels": "AB",
    "B_Assignment_Sites_Atom_Comment": "AC",
    "B_Assignment_Electronic_Types": "AF",
    "B_Assignment_Electronic_Labels": "AG",
    "B_Assignment_Electronic_Comment": "AH",
    "B_Assignment_Vibration_Types": "AK",
    "B_Assignment_Vibration_Label": "AL",
    "B_Assignment_Vibration_Bonds": "AM",
    "B_Assignment_Vibration_Comment": "AN",
    "B_Assignment_Rotation_Types": "AQ",
    "B_Assignment_Rotation_Label": "AR",
    "B_Assignment_Rotation_Comment": "AS",
    "B_Assignment_Phonon_Types": "AV",
    "B_Assignment_Phonon_Label": "AW",
    "B_Assignment_Phonon_Comment": "AX",
    "B_Assignment_Resonances_Types": "BA",
    "B_Assignment_Resonances_Band": "BB",
    "B_Assignment_Resonances_Nb": "BC",
    "B_Assignment_Resonances_Comment": "BD",
    "B_Publications_Nb": "BG",
    "B_Publications_UID": "BH",
    "B_Publications_SSHADE_Nb": "BI",
    "B_Publications_SSHADE_UID": "BJ",
    "B_Publications_Data_Nb": "BK",
    "B_Publications_Data_URL": "BL",
    "B_Publications_Comments": "BM",
    "B_Characteristics_Nb": "BP",
    "B_Characteristics_Composition": "BQ",
    "B_Characteristics_Texture": "BR",
    "B_Characteristics_T_Unit": "BW",
    "B_Characteristics_T_Value": "BT",
    "B_Characteristics_T_Error": "BU",
    "B_Characteristics_T_Formation": "BV",
    "B_Characteristics_T_Max": "BW",
    "B_Characteristics_T_Comment": "BX",
    "B_Characteristics_P_Unit": "CC",
    "B_Characteristics_P_Value": "BZ",
    "B_Characteristics_P_Error": "CA",
    "B_Characteristics_P_Formation": "CB",
    "B_Characteristics_P_Max": "CC",
    "B_Characteristics_P_Stress_type": "CD",
    "B_Characteristics_P_Comment": "CE",
    "B_Characteristics_Laser_excitation_Wavelength_Unit": "CH",
    "B_Characteristics_Laser_excitation_Wavelength": "CG",
    "B_Characteristics_Sample_Orient_mode": "CH",
    "B_Characteristics_Sample_Orient": "CI",
    "B_Characteristics_Polarization_Orient_mode": "CJ",
    "B_Characteristics_Polarization_Orient": "CK",
    "B_Characteristics_Excitation_Comment": "CL",
    "B_Characteristics_Method_Types": "CO",
    "B_Characteristics_Method_Description": "CP",
    "B_Characteristics_Method_Fit_Fct_type": "CQ",
    "B_Characteristics_Method_Fit_parameters": "CR",
    "B_Characteristics_Methods_Overlap": "CU",
    "B_Characteristics_Position_Peak_method": "CW",
    "B_Characteristics_Position_Peak": "CX",
    "B_Characteristics_Position_Peak_error": "CY",
    "B_Characteristics_Position_Center_method": "CZ",
    "B_Characteristics_Position_Center": "DA",
    "B_Characteristics_Position_Center_error": "DB",
    "B_Characteristics_Position_Evaluation": "DC",
    "B_Characteristics_Position_Comment": "DD",
    "B_Characteristics_Width_Method": "DG",
    "B_Characteristics_Width_FWHM": "DH",
    "B_Characteristics_Width_FWHM_error": "DI",
    "B_Characteristics_Width_Shape": "DJ",
    "B_Characteristics_Width_Asymm_factor": "DK",
    "B_Characteristics_Width_Asymm_factor_error": "DL",
    "B_Characteristics_Width_Evaluation": "DM",
    "B_Characteristics_Width_Comments": "DN",
    "B_Characteristics_Peak_intensity_Method": "DQ",
    "B_Characteristics_Peak_intensity_Abs_coef": "DR",
    "B_Characteristics_Peak_intensity_Abs_coef_error": "DS",
    "B_Characteristics_Peak_intensity_Abs_coef_sp": "DT",
    "B_Characteristics_Peak_intensity_Abs_coef_sp_error": "DU",
    "B_Characteristics_Peak_intensity_Relative": "DV",
    "B_Characteristics_Peak_intensity_Relative_error": "DW",
    "B_Characteristics_Peak_intensity_Strength": "DX",
    "B_Characteristics_Peak_intensity_Evaluation": "EA",
    "B_Characteristics_Peak_intensity_Comment": "EB",
    "B_Characteristics_Integrated_intensity_Method": "EE",
    "B_Characteristics_Integrated_intensity_Abs_coef": "EF",
    "B_Characteristics_Integrated_intensity_Abs_coef_error": "EG",
    "B_Characteristics_Integrated_intensity_Abs_coef_sp": "EH",
    "B_Characteristics_Integrated_intensity_Abs_coef_sp_error": "EI",
    "B_Characteristics_Integrated_intensity_Relative": "EJ",
    "B_Characteristics_Integrated_intensity_Relative_error": "EK",
    "B_Characteristics_Integrated_intensity_Strength": "EL",
    "B_Characteristics_Integrated_intensity_Evaluation": "EO",
    "B_Characteristics_Integrated_intensity_Comment": "EP",
    "B_Characteristics_Bandlist_flag": "ES"
}

# FUNCTIONS
def is_list_with_only_empty_strings(str_list):
    one_not_empty = True
    for element in str_list:
        if element != "":
            one_not_empty = False
            break
    return one_not_empty


# XLSX PARSING
def XLSX_reader(xlsx_workbook, bandlist_type):
    # VARS
    XLSX_data = dict()
    XLSX_data['BL_Import_mode'] = ''
    XLSX_data['BL_Type'] = ''
    XLSX_data['BL_Title'] = ''
    XLSX_data['BL_Description'] = ''
    XLSX_data['BL_Analysis'] = ''
    XLSX_data['BL_Global_comments'] = ''
    XLSX_data['BL_Documentation_names'] = ['']
    XLSX_data['BL_Documentation_files'] = ['']
    XLSX_data['BL_Original_data_filename'] = ''
    XLSX_data['BL_Export_filename'] = ''
    XLSX_data['BL_UID'] = ''
    XLSX_data['BL_Constituent_UID'] = ''
    XLSX_data['BL_Constituent_Primary_specie_UID'] = ''
    XLSX_data['BL_Constituent_Comments'] = ''
    XLSX_data['BL_Parents_Exp_UID'] = ['']
    XLSX_data['BL_Parents_Spectra_UID'] = ['']
    XLSX_data['BL_Parents_Comments'] = ''
    XLSX_data['BL_Spectral_Unit'] = ''
    XLSX_data['BL_Spectral_Standard'] = ''
    XLSX_data['BL_Spectral_Range_types'] = ['']
    XLSX_data['BL_Spectral_Range_min'] = ['']
    XLSX_data['BL_Spectral_Range_max'] = ['']
    XLSX_data['BL_Spectral_Comments'] = ''
    XLSX_data['BL_Spectral_Ref_pos_electronic'] = ''
    XLSX_data['BL_Spectral_Ref_pos_absorption'] = ''
    XLSX_data['BL_Validation_Quality'] = ''
    XLSX_data['BL_Validation_Date_validated'] = ''
    XLSX_data['BL_Validation_Validators_UID'] = ['']
    XLSX_data['BL_Versions_Current_version_history'] = ''
    XLSX_data['BL_Versions_Previous_version_status'] = ''
    XLSX_data['BL_Versions_Comments'] = ''
    XLSX_data['BL_Preview_x_Axis'] = ''
    XLSX_data['BL_Preview_x_Unit'] = ''
    XLSX_data['BL_Preview_x_Min'] = ''
    XLSX_data['BL_Preview_x_Max'] = ''
    XLSX_data['BL_Preview_y_Axis'] = ''
    XLSX_data['BL_Preview_y_Unit'] = ''
    XLSX_data['BL_Preview_y_Min'] = ''
    XLSX_data['BL_Preview_y_Max'] = ''
    XLSX_data["BL_Preview_y_rel_Axis"] = ''
    XLSX_data["BL_Preview_y_rel_Min"] = ''
    XLSX_data["BL_Preview_y_rel_Max"] = ''
    XLSX_data["BL_Preview_Type"] = ''
    XLSX_data['BL_Preview_Filename'] = ''
    XLSX_data['BL_Sections_Var_param'] = ''
    XLSX_data['BL_Sections_qty'] = 1
    XLSX_data['BL_Section_1_Var_param'] = ''
    XLSX_data['BL_Section_1_Title'] = ''
    XLSX_data['BL_Section_1_Description'] = ''
    XLSX_data['BL_Section_1_Bands_UID'] = ['']
    XLSX_data['BL_Section_1_Sub_sections_qty'] = 1
    XLSX_data['BL_Section_1_Sub_section_1_Title'] = ''
    XLSX_data['BL_Section_1_Sub_section_1_Description'] = ''
    XLSX_data['BL_Section_1_Sub_section_1_Bands_UID'] = ['']
    XLSX_data['B_qty'] = 1
    XLSX_data['B_1_Index'] = 1
    XLSX_data['B_1_Import_mode'] = ''
    XLSX_data['B_1_UID'] = ''
    XLSX_data['B_1_Comment'] = ''
    XLSX_data['B_1_Assignments_qty'] = 1
    XLSX_data['B_1_Assignment_1_Number'] = ''
    XLSX_data['B_1_Assignment_1_Label'] = ''
    XLSX_data['B_1_Assignment_1_Symmetry'] = ''
    XLSX_data['B_1_Assignment_1_Category'] = ''
    XLSX_data['B_1_Assignment_1_Method'] = ''
    XLSX_data['B_1_Assignment_1_Level'] = ''
    XLSX_data['B_1_Assignment_1_Evaluation'] = ''
    XLSX_data['B_1_Assignment_1_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Multiplicity_Types'] = ['']
    XLSX_data['B_1_Assignment_1_Multiplicity_Degeneracy'] = ['']
    XLSX_data['B_1_Assignment_1_Multiplicity_Other_band'] = ['']
    XLSX_data['B_1_Assignment_1_Contribution_Level'] = ''
    XLSX_data['B_1_Assignment_1_Contribution_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Transition_Species_qty'] = 1
    XLSX_data['B_1_Assignment_1_Transition_Specie_1_UID'] = ''
    XLSX_data['B_1_Assignment_1_Site_1_Molecule_labels'] = ['']
    XLSX_data['B_1_Assignment_1_Site_1_Molecule_Symm_label'] = ['']
    XLSX_data['B_1_Assignment_1_Site_1_Atom_Labels'] = ['']
    XLSX_data['B_1_Assignment_1_Site_1_Atom_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Electronic_Types'] = ['']
    XLSX_data['B_1_Assignment_1_Electronic_Labels'] = ['']
    XLSX_data['B_1_Assignment_1_Electronic_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Vibrations_qty'] = 1
    XLSX_data['B_1_Assignment_1_Vibration_1_Types'] = ''
    XLSX_data['B_1_Assignment_1_Vibration_1_Label'] = ''
    XLSX_data['B_1_Assignment_1_Vibration_1_Bonds'] = ['']
    XLSX_data['B_1_Assignment_1_Vibrations_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Rotation_Types'] = ['']
    XLSX_data['B_1_Assignment_1_Rotation_Label'] = ['']
    XLSX_data['B_1_Assignment_1_Rotation_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Phonon_Types'] = ['']
    XLSX_data['B_1_Assignment_1_Phonon_Label'] = ['']
    XLSX_data['B_1_Assignment_1_Phonon_Comment'] = ''
    XLSX_data['B_1_Assignment_1_Resonances_Types'] = ['']
    XLSX_data['B_1_Assignment_1_Resonances_Band'] = ['']
    XLSX_data['B_1_Assignment_1_Resonances_Nb'] = ['']
    XLSX_data['B_1_Assignment_1_Resonances_Comment'] = ['']
    XLSX_data['B_1_Publications_UID'] = ['']
    XLSX_data['B_1_Publications_SSHADE_UID'] = ['']
    XLSX_data['B_1_Publications_Data_URL'] = ['']
    XLSX_data['B_1_Publications_Comments'] = ''
    XLSX_data['B_1_Characteristics_qty'] = 1
    XLSX_data['B_1_Characteristic_1_Nb'] = ''
    XLSX_data['B_1_Characteristic_1_Composition'] = ''
    XLSX_data['B_1_Characteristic_1_Texture'] = ''
    XLSX_data['B_1_Characteristic_1_T_Unit'] = ''
    XLSX_data['B_1_Characteristic_1_T_Value'] = ''
    XLSX_data['B_1_Characteristic_1_T_Error'] = ''
    XLSX_data['B_1_Characteristic_1_T_Formation'] = ''
    XLSX_data['B_1_Characteristic_1_T_Max'] = ''
    XLSX_data['B_1_Characteristic_1_T_Comment'] = ''
    XLSX_data['B_1_Characteristic_1_P_Unit'] = ''
    XLSX_data['B_1_Characteristic_1_P_Value'] = ''
    XLSX_data['B_1_Characteristic_1_P_Error'] = ''
    XLSX_data['B_1_Characteristic_1_P_Formation'] = ''
    XLSX_data['B_1_Characteristic_1_P_Max'] = ''
    XLSX_data['B_1_Characteristic_1_P_Stress_type'] = ''
    XLSX_data['B_1_Characteristic_1_P_Comment'] = ''
    XLSX_data["B_1_Characteristic_1_Laser_excitation_Wavelength_Unit"] = ""
    XLSX_data['B_1_Characteristic_1_Laser_excitation_Wavelength'] = ''
    XLSX_data["B_1_Characteristic_1_Sample_Orient_mode"] = ""
    XLSX_data["B_1_Characteristic_1_Sample_Orient"] = ""
    XLSX_data["B_1_Characteristic_1_Polarization_Orient_mode"] = ""
    XLSX_data["B_1_Characteristic_1_Polarization_Orient"] = ""
    XLSX_data["B_1_Characteristic_1_Excitation_Comment"] = ""
    XLSX_data['B_1_Characteristic_1_Methods_qty'] = 1
    XLSX_data['B_1_Characteristic_1_Method_1_Types'] = ''
    XLSX_data['B_1_Characteristic_1_Method_1_Description'] = ''
    XLSX_data['B_1_Characteristic_1_Method_1_Fit_Fct_type'] = ''
    XLSX_data['B_1_Characteristic_1_Method_1_Fit_parameters'] = ''
    XLSX_data['B_1_Characteristic_1_Methods_Overlap'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Peak_method'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Peak'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Peak_error'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Center_method'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Center'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Center_error'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Evaluation'] = ''
    XLSX_data['B_1_Characteristic_1_Position_Comment'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Method'] = ''
    XLSX_data['B_1_Characteristic_1_Width_FWHM'] = ''
    XLSX_data['B_1_Characteristic_1_Width_FWHM_error'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Shape'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Asymm_factor'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Asymm_factor_error'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Evaluation'] = ''
    XLSX_data['B_1_Characteristic_1_Width_Comments'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Method'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Abs_coef'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_error'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_sp'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_sp_error'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Relative'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Relative_error'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Strength'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Evaluation'] = ''
    XLSX_data['B_1_Characteristic_1_Peak_intensity_Comment'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Method'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_error'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_sp'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_sp_error'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Relative'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Relative_error'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Strength'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Evaluation'] = ''
    XLSX_data['B_1_Characteristic_1_Integrated_intensity_Comment'] = ''
    XLSX_data['B_1_Characteristic_1_Bandlist_flag'] = ''
    position_data = dict()
    position_data['BL_Import_mode'] = ('', '', '')
    position_data['BL_Type'] = ('', '', '')
    position_data['BL_Title'] = ('', '', '')
    position_data['BL_Description'] = ('', '', '')
    position_data['BL_Analysis'] = ('', '', '')
    position_data['BL_Global_comments'] = ('', '', '')
    position_data['BL_Documentation_names'] = ('', '', '')
    position_data['BL_Documentation_files'] = ('', '', '')
    position_data['BL_Original_data_filename'] = ('', '', '')
    position_data['BL_Export_filename'] = ('', '', '')
    position_data['BL_UID'] = ('', '', '')
    position_data['BL_Constituent_UID'] = ('', '', '')
    position_data['BL_Constituent_Primary_specie_UID'] = ('', '', '')
    position_data['BL_Constituent_Comments'] = ('', '', '')
    position_data['BL_Parents_Exp_UID'] = ('', '', '')
    position_data['BL_Parents_Spectra_UID'] = ('', '', '')
    position_data['BL_Parents_Comments'] = ('', '', '')
    position_data['BL_Spectral_Unit'] = ('', '', '')
    position_data['BL_Spectral_Standard'] = ('', '', '')
    position_data['BL_Spectral_Range_types'] = ('', '', '')
    position_data['BL_Spectral_Range_min'] = ('', '', '')
    position_data['BL_Spectral_Range_max'] = ('', '', '')
    position_data['BL_Spectral_Comments'] = ('', '', '')
    position_data['BL_Spectral_Ref_pos_electronic'] = ('', '', '')
    position_data['BL_Spectral_Ref_pos_absorption'] = ('', '', '')
    position_data['BL_Validation_Quality'] = ('', '', '')
    position_data['BL_Validation_Date_validated'] = ('', '', '')
    position_data['BL_Validation_Validators_UID'] = ('', '', '')
    position_data['BL_Versions_Current_version_history'] = ('', '', '')
    position_data['BL_Versions_Previous_version_status'] = ('', '', '')
    position_data['BL_Versions_Comments'] = ('', '', '')
    position_data['BL_Preview_x_Axis'] = ('', '', '')
    position_data['BL_Preview_x_Unit'] = ('', '', '')
    position_data['BL_Preview_x_Min'] = ('', '', '')
    position_data['BL_Preview_x_Max'] = ('', '', '')
    position_data['BL_Preview_y_Axis'] = ('', '', '')
    position_data['BL_Preview_y_Unit'] = ('', '', '')
    position_data['BL_Preview_y_Min'] = ('', '', '')
    position_data['BL_Preview_y_Max'] = ('', '', '')
    position_data["BL_Preview_y_rel_Axis"] = ('', '', '')
    position_data["BL_Preview_y_rel_Min"] = ('', '', '')
    position_data["BL_Preview_y_rel_Max"] = ('', '', '')
    position_data['BL_Preview_Filename'] = ('', '', '')
    position_data['BL_Sections_Var_param'] = ('', '', '')
    position_data['BL_Section_1_Var_param'] = ('', '', '')
    position_data['BL_Section_1_Title'] = ('', '', '')
    position_data['BL_Section_1_Description'] = ('', '', '')
    position_data['BL_Section_1_Bands_UID'] = ('', '', '')
    position_data['BL_Section_1_Sub_section_1_Title'] = ('', '', '')
    position_data['BL_Section_1_Sub_section_1_Description'] = ('', '', '')
    position_data['BL_Section_1_Sub_section_1_Bands_UID'] = ('', '', '')
    position_data['B_1_Index'] = ('', '', '')
    position_data['B_1_Import_mode'] = ('', '', '')
    position_data['B_1_UID'] = ('', '', '')
    position_data['B_1_Comment'] = ('', '', '')
    position_data['B_1_Publications_Comments'] = ('', '', '')
    position_data['B_1_Publications_UID'] = ('', '', '')
    position_data['B_1_Publications_SSHADE_UID'] = ('', '', '')
    position_data['B_1_Assignment_1_Number'] = ('', '', '')
    position_data['B_1_Assignment_1_Label'] = ('', '', '')
    position_data['B_1_Assignment_1_Category'] = ('', '', '')
    position_data['B_1_Assignment_1_Method'] = ('', '', '')
    position_data['B_1_Assignment_1_Level'] = ('', '', '')
    position_data['B_1_Assignment_1_Symmetry'] = ('', '', '')
    position_data['B_1_Assignment_1_Evaluation'] = ('', '', '')
    position_data['B_1_Assignment_1_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Multiplicity_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Multiplicity_Degeneracy'] = ('', '', '')
    position_data['B_1_Assignment_1_Multiplicity_Other_band'] = ('', '', '')
    position_data['B_1_Assignment_1_Contribution_Level'] = ('', '', '')
    position_data['B_1_Assignment_1_Contribution_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Transition_Specie_1_UID'] = ('', '', '')
    position_data['B_1_Assignment_1_Site_1_Molecule_labels'] = ('', '', '')
    position_data['B_1_Assignment_1_Site_1_Molecule_Symm_label'] = ('', '', '')
    position_data['B_1_Assignment_1_Site_1_Atom_Labels'] = ('', '', '')
    position_data['B_1_Assignment_1_Site_1_Atom_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Electronic_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Electronic_Labels'] = ('', '', '')
    position_data['B_1_Assignment_1_Electronic_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Vibration_1_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Vibration_1_Label'] = ('', '', '')
    position_data['B_1_Assignment_1_Vibration_1_Bonds'] = ('', '', '')
    position_data['B_1_Assignment_1_Vibrations_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Rotation_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Rotation_Label'] = ('', '', '')
    position_data['B_1_Assignment_1_Rotation_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Phonon_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Phonon_Label'] = ('', '', '')
    position_data['B_1_Assignment_1_Phonon_Comment'] = ('', '', '')
    position_data['B_1_Assignment_1_Resonances_Types'] = ('', '', '')
    position_data['B_1_Assignment_1_Resonances_Band'] = ('', '', '')
    position_data['B_1_Assignment_1_Resonances_Nb'] = ('', '', '')
    position_data['B_1_Assignment_1_Resonances_Comment'] = ('', '', '')
    position_data['B_1_Characteristic_1_Nb'] = ('', '', '')
    position_data['B_1_Characteristic_1_Composition'] = ('', '', '')
    position_data['B_1_Characteristic_1_Texture'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Unit'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Value'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Error'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Formation'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Max'] = ('', '', '')
    position_data['B_1_Characteristic_1_T_Comment'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Unit'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Value'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Error'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Formation'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Max'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Stress_type'] = ('', '', '')
    position_data['B_1_Characteristic_1_P_Comment'] = ('', '', '')
    position_data["B_1_Characteristic_1_Laser_excitation_Wavelength_Unit"] = ('', '', '')
    position_data['B_1_Characteristic_1_Laser_excitation_Wavelength'] = ('', '', '')
    position_data["B_1_Characteristic_1_Sample_Orient_mode"] = ('', '', '')
    position_data["B_1_Characteristic_1_Sample_Orient"] = ('', '', '')
    position_data["B_1_Characteristic_1_Polarization_Orient_mode"] = ('', '', '')
    position_data["B_1_Characteristic_1_Polarization_Orient"] = ('', '', '')
    position_data["B_1_Characteristic_1_Excitation_Comment"] = ('', '', '')
    position_data['B_1_Characteristic_1_Method_1_Types'] = ('', '', '')
    position_data['B_1_Characteristic_1_Method_1_Description'] = ('', '', '')
    position_data['B_1_Characteristic_1_Method_1_Fit_Fct_type'] = ('', '', '')
    position_data['B_1_Characteristic_1_Method_1_Fit_parameters'] = ('', '', '')
    position_data['B_1_Characteristic_1_Methods_Overlap'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Peak_method'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Peak'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Peak_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Center_method'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Center'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Center_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Evaluation'] = ('', '', '')
    position_data['B_1_Characteristic_1_Position_Comment'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Method'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_FWHM'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_FWHM_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Shape'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Asymm_factor'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Asymm_factor_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Evaluation'] = ('', '', '')
    position_data['B_1_Characteristic_1_Width_Comments'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Method'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Abs_coef'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_sp'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Abs_coef_sp_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Relative'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Relative_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Strength'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Evaluation'] = ('', '', '')
    position_data['B_1_Characteristic_1_Peak_intensity_Comment'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Method'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_sp'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Abs_coef_sp_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Relative'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Relative_error'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Strength'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Evaluation'] = ('', '', '')
    position_data['B_1_Characteristic_1_Integrated_intensity_Comment'] = ('', '', '')
    position_data['B_1_Characteristic_1_Bandlist_flag'] = ('', '', '')
    workbook = openpyxl.load_workbook(xlsx_workbook, data_only=True, read_only=False)
    # data_only: do not take into account formula, only data
    # read_only=True economizes the physical memory and therefore it is very slow

    # INNER FUNCTIONS
    # function searches for an element with case_ID in sheet
    def attribution(sheet, case_ID, no_select_one=True):
        """
        This function returns value situated in the 'case_ID' in a given 'sheet'.
        If the value is None then it returns "".
        If the value is "Select one" then it can return "" or "Select one" depending on no_select_one value.
        If the value is a date or time then it returns it in date Python format.
        If the value is a string then it cleans it out of white spaces and returns.
        In all other cases it returns the 'case_ID' value.
        """
        if sheet[case_ID].value is not None:
            if type(sheet[case_ID].value) == str:
                if sheet[case_ID].value.strip() == "Select one":
                    if no_select_one:
                        return ""
                    else:
                        return "Select one"
                else:
                    return sheet[case_ID].value.strip()
            else:
                try:
                    return sheet[case_ID].value.strftime("%Y-%m-%d")
                except:
                    return str(sheet[case_ID].value)
        else:
            return ""

    # function to collect multiples
    def multiple_collect(sheet, column_letter, line_start, line_stop, empty_lines=False):
        """
        This function returns an array made of values in 'sheet' situated in column 'column_letter'
        within lines line_start and line_stop (both included).
        If empty_lines is False (by default), it will returns only non empty values. Otherwise it will returns all values.
        If there is no values found, it returns [""] - an array with one empty string.
        """
        temp_array = []
        for index in range(line_start, line_stop + 1):
            if not empty_lines:
                if attribution(sheet, f'{column_letter}{index}'):
                    temp_array.append(attribution(sheet, f'{column_letter}{index}'))
            else:
                temp_array.append(attribution(sheet, f'{column_letter}{index}'))
        if temp_array:
            return temp_array
        else:
            return [""]

    # function to collect positions of multiples
    def multiple_position(sheet, column_letter, line_start, line_stop, empty_lines=False):
        """
        This function takes a 'sheet', a column 'column_letter' and lines line_start and line_stop (both included).
        It returns a tuple with 'column_letter', line_start and an index with the last non-empty line number.
        """
        if column_letter == f"{col_n['B_Assignment_Rotation_Label']}":
            a = 0
        if empty_lines:
            return column_letter, line_start, line_stop
        last_line = line_start
        for index in range(line_start, line_stop + 1):
            if attribution(sheet, f'{column_letter}{index}'):
                last_line = index
            if not attribution(sheet, f'{column_letter}{index}') and index == line_start:
                last_line = last_line - 1
        if line_start <= last_line:
            return column_letter, line_start, last_line
        else:
            return "", "", ""

    # function to collect ordered multiples
    def multiple_ordered_collect(sheet, data_column_letter, order_column_letter, line_start, line_stop):
        """
        This function takes two sets: a column with data and a column with order indexes (int numbers).
        Both are in 'sheet' within line_start and line_stop (included).
        It returns an array with the data ordered by order index.
        If there is no values then it returns [""] - an array with one empty string.
        """
        temp_array_1 = []
        for index in range(line_start, line_stop + 1):
            if attribution(sheet, f'{data_column_letter}{index}') and attribution(sheet, f'{order_column_letter}{index}'):
                temp_array_1.append((attribution(sheet, f'{order_column_letter}{index}'), attribution(sheet, f'{data_column_letter}{index}')))
                try:
                    int(attribution(sheet, f'{order_column_letter}{index}'))
                except ValueError:
                    return f'There is a letter or other non-int symbol in Band, {order_column_letter}{index}'
                except Exception as e:
                    return f'Other error in multiple_ordered_collect function: {str(e)}.'
        temp_array_1.sort(key=lambda x: int(x[0]))
        temp_array_2 = []
        for item in temp_array_1:
            temp_array_2.append(item[1])
        del temp_array_1
        if temp_array_2:
            return temp_array_2
        else:
            return [""]

    # function to collect borders of sub_multiples
    def sub_multiples(sheet, column_letter_list, line_start, line_stop):
        """
        This function takes a 'sheet', a column letter and line_start and line_stop (both included).
        It searches for positions of all non-empty elements (sub_multiples).
        It returns an array of tuples with start and stop line numbers for these elements.
        """
        all_starts = []
        new_line = False
        for index in range(line_start, line_stop + 1):
            for column_letter in column_letter_list:
                if new_line or attribution(sheet, f'{column_letter}{index}'):
                    new_line = True
            if new_line:
                all_starts.append(index)
                new_line = False
        all_starts.append(line_stop)
        selected_items = []
        if len(all_starts) > 1:
            for index in range(0, len(all_starts) - 1):
                if index == len(all_starts) - 2:
                    selected_items.append((all_starts[index], all_starts[index + 1]))
                else:
                    selected_items.append((all_starts[index], all_starts[index + 1] - 1))
        return selected_items

    # function to collect borders of sub_multiples with order
    def sub_multiples_ordered(sheet, column_letter_order_list, column_letter_select, line_start, line_stop):
        """
        This function takes two sets: the data in column_letter_select and its order indexes (int numbers) in column_letter_order.
        Both are in a 'sheet', within line_start and line_stop (both included).
        It searches for positions of all non-empty elements (sub_multiples).
        It returns an array of tuples with an order index and with start and stop line numbers for these elements.
        Elements of this array are ordered by the order index.
        """
        all_starts = []
        new_line = False
        for index in range(line_start, line_stop + 1):
            for column_letter in column_letter_order_list:
                if new_line or attribution(sheet, f'{column_letter}{index}'):
                    new_line = True
            if new_line:
                all_starts.append(index)
                new_line = False
        all_starts.append(line_stop)
        selected_items = []
        for index in range(0, len(all_starts) - 1):
            if attribution(sheet, f'{column_letter_select}{all_starts[index]}'):
                if index == len(all_starts) - 2:
                    selected_items.append((attribution(sheet, f'{column_letter_select}{all_starts[index]}'), all_starts[index], all_starts[index + 1]))
                else:
                    selected_items.append((attribution(sheet, f'{column_letter_select}{all_starts[index]}'), all_starts[index], all_starts[index + 1] - 1))
                try:
                    int(attribution(sheet, f'{column_letter_select}{all_starts[index]}'))
                except ValueError:
                    return f'There is a letter or other non-int symbol in Band, {column_letter_select}{all_starts[index]}'
                except Exception as e:
                    return f'Other error in sub_multiples function: {str(e)}.'
        selected_items.sort(key=lambda x: int(x[0]))
        return selected_items

    # workbook PARSE and READ
    for sheet in workbook:
        # Bandlist sheet data
        if sheet.title == 'Bandlist_' + bandlist_type:
            # Bandlist
            XLSX_data["BL_Import_mode"] = attribution(sheet, f'C3')
            position_data["BL_Import_mode"] = (f'C', 3, 3)
            XLSX_data["BL_Type"] = attribution(sheet, f'C4')
            position_data["BL_Type"] = (f'C', 4, 4)
            XLSX_data["BL_Title"] = attribution(sheet, f'C5')
            position_data["BL_Title"] = (f'C', 5, 5)
            XLSX_data["BL_Description"] = attribution(sheet, f'C6')
            position_data["BL_Description"] = (f'C', 6, 6)
            XLSX_data["BL_Analysis"] = attribution(sheet, f'C7')
            position_data["BL_Analysis"] = (f'C', 7, 7)
            XLSX_data["BL_Global_comments"] = attribution(sheet, f'C8')
            position_data["BL_Global_comments"] = (f'C', 8, 8)
            # Files
            XLSX_data["BL_Documentation_names"] = multiple_collect(sheet, f'C', 9, 11, True)
            position_data["BL_Documentation_names"] = (f'C', 9, 11)
            XLSX_data["BL_Documentation_files"] = multiple_collect(sheet, f'E', 9, 11, True)
            position_data["BL_Documentation_files"] = (f'E', 9, 11)
            XLSX_data["BL_Original_data_filename"] = attribution(sheet, f'C12')
            position_data["BL_Original_data_filename"] = (f'C', 12, 12)
            XLSX_data["BL_Export_filename"] = attribution(sheet, f'C13')
            position_data["BL_Export_filename"] = (f'C', 13, 13)
            # UIDs
            XLSX_data["BL_UID"] = attribution(sheet, f'A17')
            position_data["BL_UID"] = (f'A', 17, 17)
            XLSX_data["BL_Constituent_UID"] = attribution(sheet, f'B17')
            position_data["BL_Constituent_UID"] = (f'B', 17, 17)
            XLSX_data["BL_Constituent_Primary_specie_UID"] = attribution(sheet, f'C17')
            position_data["BL_Constituent_Primary_specie_UID"] = (f'C', 17, 17)
            XLSX_data["BL_Constituent_Comments"] = attribution(sheet, f'D17')
            position_data["BL_Constituent_Comments"] = (f'D', 17, 17)
            # Parents
            XLSX_data["BL_Parents_Exp_UID"] = multiple_collect(sheet, f'A', 25, 28, True)
            position_data["BL_Parents_Exp_UID"] = (f'A', 25, 28)
            XLSX_data["BL_Parents_Spectra_UID"] = multiple_collect(sheet, f'B', 25, 28, True)
            position_data["BL_Parents_Spectra_UID"] = (f'B', 25, 28)
            XLSX_data["BL_Parents_Comments"] = attribution(sheet, f'C25')
            position_data["BL_Parents_Comments"] = (f'C', 25, 25)
            # Spectral
            XLSX_data["BL_Spectral_Unit"] = attribution(sheet, f'A34')
            position_data["BL_Spectral_Unit"] = (f'A', 34, 34)
            XLSX_data["BL_Spectral_Standard"] = attribution(sheet, f'B34')
            position_data["BL_Spectral_Standard"] = (f'B', 34, 34)
            XLSX_data["BL_Spectral_Range_types"] = multiple_collect(sheet, f'C', 34, 39)
            position_data["BL_Spectral_Range_types"] = (f'C', 34, 39)
            XLSX_data["BL_Spectral_Range_min"] = multiple_collect(sheet, f'D', 34, 37, True)
            position_data["BL_Spectral_Range_min"] = (f'D', 34, 37)
            XLSX_data["BL_Spectral_Range_max"] = multiple_collect(sheet, f'E', 34, 37, True)
            position_data["BL_Spectral_Range_max"] = (f'E', 34, 37)
            XLSX_data["BL_Spectral_Comments"] = attribution(sheet, f'F34')
            position_data["BL_Spectral_Comments"] = (f'F', 34, 34)
            XLSX_data["BL_Spectral_Ref_pos_electronic"] = attribution(sheet, f'G34')
            position_data["BL_Spectral_Ref_pos_electronic"] = (f'G', 34, 34)
            XLSX_data["BL_Spectral_Ref_pos_absorption"] = attribution(sheet, f'H34')
            position_data["BL_Spectral_Ref_pos_absorption"] = (f'H', 34, 34)
            # Validation
            XLSX_data["BL_Validation_Quality"] = attribution(sheet, f'A44')
            position_data["BL_Validation_Quality"] = (f'A', 44, 44)
            XLSX_data["BL_Validation_Date_validated"] = attribution(sheet, f'B44')
            position_data["BL_Validation_Date_validated"] = (f'B', 44, 44)
            XLSX_data["BL_Validation_Validators_UID"] = multiple_collect(sheet, f'C', 44, 47)
            position_data["BL_Validation_Validators_UID"] = (f'C', 44, 47)
            # Versions
            XLSX_data["BL_Versions_Current_version_history"] = attribution(sheet, f'A53')
            position_data["BL_Versions_Current_version_history"] = (f'A', 53, 53)
            XLSX_data["BL_Versions_Previous_version_status"] = attribution(sheet, f'B53')
            position_data["BL_Versions_Previous_version_status"] = (f'B', 53, 53)
            XLSX_data["BL_Versions_Comments"] = attribution(sheet, f'C53')
            position_data["BL_Versions_Comments"] = (f'C', 53, 53)
            # Preview
            XLSX_data["BL_Preview_x_Axis"] = attribution(sheet, f'B57')
            position_data["BL_Preview_x_Axis"] = (f'B', 57, 57)
            XLSX_data["BL_Preview_x_Unit"] = attribution(sheet, f'C57')
            position_data["BL_Preview_x_Unit"] = (f'C', 57, 57)
            XLSX_data["BL_Preview_x_Min"] = attribution(sheet, f'D57')
            position_data["BL_Preview_x_Min"] = (f'D', 57, 57)
            XLSX_data["BL_Preview_x_Max"] = attribution(sheet, f'E57')
            position_data["BL_Preview_x_Max"] = (f'E', 57, 57)
            XLSX_data["BL_Preview_y_Axis"] = attribution(sheet, f'B58')
            position_data["BL_Preview_y_Axis"] = (f'B', 58, 58)
            XLSX_data["BL_Preview_y_Unit"] = attribution(sheet, f'C58')
            position_data["BL_Preview_y_Unit"] = (f'C', 58, 58)
            XLSX_data["BL_Preview_y_Min"] = attribution(sheet, f'D58')
            position_data["BL_Preview_y_Min"] = (f'D', 58, 58)
            XLSX_data["BL_Preview_y_Max"] = attribution(sheet, f'E58')
            position_data["BL_Preview_y_Max"] = (f'E', 58, 58)
            XLSX_data["BL_Preview_y_rel_Axis"] = attribution(sheet, f'B59')
            position_data["BL_Preview_y_rel_Axis"] = (f'B', 59, 59)
            XLSX_data["BL_Preview_y_rel_Min"] = attribution(sheet, f'D59')
            position_data["BL_Preview_y_rel_Min"] = (f'D', 59, 59)
            XLSX_data["BL_Preview_y_rel_Max"] = attribution(sheet, f'E59')
            position_data["BL_Preview_y_rel_Max"] = (f'E', 59, 59)
            XLSX_data["BL_Preview_Type"] = attribution(sheet, f'E60')
            position_data["BL_Preview_Type"] = (f'E', 60, 60)
            XLSX_data["BL_Preview_Filename"] = attribution(sheet, f'B60')
            position_data["BL_Preview_Filename"] = (f'B', 60, 60)
            # Structure
            XLSX_data["BL_Sections_Var_param"] = attribution(sheet, f'C64')
            position_data["BL_Sections_Var_param"] = (f'C', 64, 64)
            # Structure : search for section starts
            section_start = 66
            section_list = [section_start]
            a_index = section_start
            section_end = False
            while attribution(sheet, f'A{a_index}') != 'END' and a_index < sheet.max_row:
                if attribution(sheet, f'A{a_index}') != '':
                    if section_end:
                        section_list.append(a_index)
                        section_end = False
                else:
                    if not section_end:
                        section_end = True
                a_index = a_index + 1
            if a_index != section_start:
                section_list.append(a_index)
            # Structure : sections read
            position_data["BL_Section_1_Var_param"] = (f'C', 66, 66)
            position_data["BL_Section_1_Title"] = (f'B', 67, 67)
            position_data["BL_Section_1_Description"] = (f'B', 68, 68)
            if len(section_list) > 1:
                XLSX_data["BL_Sections_qty"] = len(section_list) - 1
                for i in range(1, len(section_list)):
                    XLSX_data[f"BL_Section_{i}_Var_param"] = attribution(sheet, f'C{section_list[i - 1]}')
                    position_data[f"BL_Section_{i}_Var_param"] = (f'C', section_list[i - 1], section_list[i - 1])
                    XLSX_data[f"BL_Section_{i}_Title"] = attribution(sheet, f'B{section_list[i - 1] + 1}')
                    position_data[f"BL_Section_{i}_Title"] = (f'B', section_list[i - 1] + 1, section_list[i - 1] + 1)
                    XLSX_data[f"BL_Section_{i}_Description"] = attribution(sheet, f'B{section_list[i - 1] + 2}')
                    position_data[f"BL_Section_{i}_Description"] = (f'B', section_list[i - 1] + 2, section_list[i - 1] + 2)
                    XLSX_data[f"BL_Section_{i}_Bands_UID"] = multiple_collect(sheet, f'B', section_list[i - 1] + 3, section_list[i] - 1)
                    position_data[f"BL_Section_{i}_Bands_UID"] = multiple_position(sheet, f'B', section_list[i - 1] + 3, section_list[i] - 1)
                    sub_section_index = 1
                    check_empty = 1
                    while check_empty:
                        band_exist = False
                        for band_index in range(section_list[i - 1] + 3, section_list[i]):
                            if attribution(sheet, f'{letters[sub_section_index + 2]}{band_index}'):
                                band_exist = True
                                break
                        if attribution(sheet, f'{letters[sub_section_index + 2]}{section_list[i - 1] + 1}') or attribution(sheet, f'{letters[sub_section_index + 2]}{section_list[i - 1] + 2}') or band_exist:
                            XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Title"] = attribution(sheet, f'{letters[sub_section_index + 2]}{section_list[i - 1] + 1}')
                            position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Title"] = (letters[sub_section_index + 2], section_list[i - 1] + 1, section_list[i - 1] + 1)
                            XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Description"] = attribution(sheet, f'{letters[sub_section_index + 2]}{section_list[i - 1] + 2}')
                            position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Description"] = (letters[sub_section_index + 2], section_list[i - 1] + 2, section_list[i - 1] + 2)
                            XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Bands_UID"] = multiple_collect(sheet, letters[sub_section_index + 2], section_list[i - 1] + 3, section_list[i] - 1)
                            position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Bands_UID"] = multiple_position(sheet, letters[sub_section_index + 2], section_list[i - 1] + 3, section_list[i] - 1)
                            sub_section_index = sub_section_index + 1
                        else:
                            check_empty = 0
                            if sub_section_index == 1:
                                XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Title"] = ""
                                position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Title"] = ('', '', '')
                                XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Description"] = ""
                                position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Description"] = ('', '', '')
                                XLSX_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Bands_UID"] = [""]
                                position_data[f"BL_Section_{i}_Sub_section_{sub_section_index}_Bands_UID"] = ('', '', '')
                    XLSX_data[f"BL_Section_{i}_Sub_sections_qty"] = sub_section_index - 1
        # Bands sheet data
        if sheet.title == 'Bands':
            # search for bands
            start_line = 0
            band_start = False
            good_band = False
            first_time = True
            B_Index = 1
            band_list = []
            for B_Index in range(1, sheet.max_row + 1):
                if attribution(sheet, f'B{B_Index}') == 'END':
                    break
                if attribution(sheet, f'B{B_Index}') == 'START':
                    start_line = B_Index
                    band_start = True
                if band_start and attribution(sheet, f'{col_n["B_Import_mode"]}{B_Index}') != "":
                    if attribution(sheet, f'{col_n["B_Index"]}{B_Index}') and attribution(sheet, f'{col_n["B_Type"]}{B_Index}') == attribution(workbook[f"Bandlist_{bandlist_type}"], f"C4"):
                        if not first_time and good_band:
                            try:
                                int(band_index)
                            except ValueError:
                                raise Exception(f'There is a letter or other non-int symbol in Band, A{B_Index}')
                            except Exception as e:
                                raise Exception(f'Other error during band selection: {str(e)}.')
                            band_list.append((band_index, band_begin, B_Index - 1))
                        band_index = attribution(sheet, f'A{B_Index}')
                        band_begin = B_Index
                        first_time = False
                        good_band = True
                    else:
                        if not first_time and good_band:
                            try:
                                int(band_index)
                            except ValueError:
                                raise Exception(f'There is a letter or other non-int symbol in Band, A{B_Index}')
                            except Exception as e:
                                raise Exception(f'Other error during band selection: {str(e)}.')
                            band_list.append((band_index, band_begin, B_Index - 1))
                        good_band = False
            if not first_time and good_band:
                try:
                    int(band_index)
                except ValueError:
                    raise Exception(f'There is a letter or other non-int symbol in Band, A{B_Index}')
                except Exception as e:
                    raise Exception(f'Other error during band selection: {str(e)}.')
                band_list.append((band_index, band_begin, B_Index))
            # order the selected bands
            band_list.sort(key=lambda x: int(x[0]))
            # Bands READ
            if len(band_list) != 0:
                XLSX_data["B_qty"] = len(band_list)
            for index_lvl_1 in range(1, len(band_list) + 1):
                # Bands
                XLSX_data[f'B_{index_lvl_1}_Index'] = attribution(sheet, f"{col_n['B_Index']}{band_list[index_lvl_1 - 1][1]}")
                position_data[f'B_{index_lvl_1}_Index'] = (f"{col_n['B_Index']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                XLSX_data[f"B_{index_lvl_1}_Import_mode"] = attribution(sheet, f"{col_n['B_Import_mode']}{band_list[index_lvl_1 - 1][1]}")
                position_data[f"B_{index_lvl_1}_Import_mode"] = (f"{col_n['B_Import_mode']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                XLSX_data[f"B_{index_lvl_1}_UID"] = attribution(sheet, f"{col_n['B_UID']}{band_list[index_lvl_1 - 1][1]}")
                position_data[f"B_{index_lvl_1}_UID"] = (f"{col_n['B_UID']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                XLSX_data[f"B_{index_lvl_1}_Comment"] = attribution(sheet, f"{col_n['B_Comment']}{band_list[index_lvl_1 - 1][1]}")
                position_data[f"B_{index_lvl_1}_Comment"] = (f"{col_n['B_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                # Assignment
                temp_result = sub_multiples_ordered(sheet, [f"{col_n['B_Assignment_Label']}", f"{col_n['B_Assignment_Number']}"], f"{col_n['B_Assignment_Number']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if type(temp_result) == str:
                    raise Exception(temp_result)
                assignment_list = temp_result
                if len(assignment_list) == 0:
                    XLSX_data[f"B_{index_lvl_1}_Assignments_qty"] = 1
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Number"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Number"] = (f'{col_n["B_Assignment_Number"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Label"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Label"] = (f'{col_n["B_Assignment_Label"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Category"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Category"] = (f'{col_n["B_Assignment_Category"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Method"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Method"] = (f'{col_n["B_Assignment_Method"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Level"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Level"] = (f'{col_n["B_Assignment_Level"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Symmetry"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Symmetry"] = (f'{col_n["B_Assignment_Symmetry"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Evaluation"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Evaluation"] = (f'{col_n["B_Assignment_Evaluation"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Comment"] = (f'{col_n["B_Assignment_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Types"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Types"] = (f'{col_n["B_Assignment_Multiplicity_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Degeneracy"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Degeneracy"] = (f'{col_n["B_Assignment_Multiplicity_Degeneracy"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Other_band"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Multiplicity_Other_band"] = (f'{col_n["B_Assignment_Multiplicity_Other_band"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Contribution_Level"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Contribution_Level"] = (f'{col_n["B_Assignment_Contribution_Level"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Contribution_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Contribution_Comment"] = (f'{col_n["B_Assignment_Contribution_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : primary_species
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Transition_Species_qty"] = 1
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Transition_Specie_1_UID"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Transition_Specie_1_UID"] = (f'{col_n["B_Assignment_Transition_Specie_UID"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Molecule_labels"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Molecule_labels"] = (f'{col_n["B_Assignment_Sites_Molecule_labels"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Molecule_Symm_label"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Molecule_Symm_label"] = (f'{col_n["B_Assignment_Sites_Molecule_Symm_label"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Atom_Labels"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Atom_Labels"] = (f'{col_n["B_Assignment_Sites_Atom_Labels"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Atom_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Site_1_Atom_Comment"] = (f'{col_n["B_Assignment_Sites_Atom_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : electronic_modes
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Types"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Types"] = (f'{col_n["B_Assignment_Electronic_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Labels"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Labels"] = (f'{col_n["B_Assignment_Electronic_Labels"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Electronic_Comment"] = (f'{col_n["B_Assignment_Electronic_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : vibration_modes
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Vibrations_qty"] = 1
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Types"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Types"] = (f'{col_n["B_Assignment_Vibration_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Label"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Label"] = (f'{col_n["B_Assignment_Vibration_Label"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Bonds"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Vibration_1_Bonds"] = (f'{col_n["B_Assignment_Vibration_Bonds"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Vibrations_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Vibrations_Comment"] = (f'{col_n["B_Assignment_Vibration_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : rotation_modes
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Types"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Types"] = (f'{col_n["B_Assignment_Rotation_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Label"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Label"] = (f'{col_n["B_Assignment_Rotation_Label"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Rotation_Comment"] = (f'{col_n["B_Assignment_Rotation_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : phonon_modes
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Types"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Types"] = (f'{col_n["B_Assignment_Phonon_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Label"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Label"] = (f'{col_n["B_Assignment_Phonon_Label"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Assignment_1_Phonon_Comment"] = (f'{col_n["B_Assignment_Phonon_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Transition : resonances
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Types"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Types"] = (f'{col_n["B_Assignment_Resonances_Types"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Band"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Band"] = (f'{col_n["B_Assignment_Resonances_Band"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Nb"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Nb"] = (f'{col_n["B_Assignment_Resonances_Nb"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Comment"] = [""]
                    position_data[f"B_{index_lvl_1}_Assignment_1_Resonances_Comment"] = (f'{col_n["B_Assignment_Resonances_Comment"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                else:
                    XLSX_data[f"B_{index_lvl_1}_Assignments_qty"] = len(assignment_list)
                    for index_lvl_2 in range(1, len(assignment_list) + 1):
                        # Assignment: general
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Number"] = attribution(sheet, f'{col_n["B_Assignment_Number"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Number"] = (f'{col_n["B_Assignment_Number"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Label"] = attribution(sheet, f'{col_n["B_Assignment_Label"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Label"] = (f'{col_n["B_Assignment_Label"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Category"] = attribution(sheet, f'{col_n["B_Assignment_Category"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Category"] = (f'{col_n["B_Assignment_Category"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Method"] = attribution(sheet, f'{col_n["B_Assignment_Method"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Method"] = (f'{col_n["B_Assignment_Method"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Level"] = attribution(sheet, f'{col_n["B_Assignment_Level"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Level"] = (f'{col_n["B_Assignment_Level"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Symmetry"] = attribution(sheet, f'{col_n["B_Assignment_Symmetry"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Symmetry"] = (f'{col_n["B_Assignment_Symmetry"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Evaluation"] = attribution(sheet, f'{col_n["B_Assignment_Evaluation"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Evaluation"] = (f'{col_n["B_Assignment_Evaluation"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Comment"] = (f'{col_n["B_Assignment_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Multiplicity
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Types"] = multiple_collect(sheet, f'{col_n["B_Assignment_Multiplicity_Types"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], empty_lines=True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Types"] = (f'{col_n["B_Assignment_Multiplicity_Types"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Degeneracy"] = multiple_collect(sheet, f'{col_n["B_Assignment_Multiplicity_Degeneracy"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], empty_lines=True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Degeneracy"] = (f'{col_n["B_Assignment_Multiplicity_Degeneracy"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Other_band"] = multiple_collect(sheet, f'{col_n["B_Assignment_Multiplicity_Other_band"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], empty_lines=True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Multiplicity_Other_band"] = (f'{col_n["B_Assignment_Multiplicity_Other_band"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Contribution_Level"] = attribution(sheet, f'{col_n["B_Assignment_Contribution_Level"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Contribution_Level"] = (f'{col_n["B_Assignment_Contribution_Level"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Contribution_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Contribution_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Contribution_Comment"] = (f'{col_n["B_Assignment_Contribution_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Transition : primary_species
                        selected_items = sub_multiples(sheet, [f"{col_n['B_Assignment_Transition_Specie_UID']}"], assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        if len(selected_items) == 0:
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Species_qty"] = 1
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Specie_1_UID"] = ""
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Specie_1_UID"] = (f'{col_n["B_Assignment_Transition_Specie_UID"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Molecule_labels"] = [""]
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Molecule_labels"] = (f'{col_n["B_Assignment_Sites_Molecule_labels"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Molecule_Symm_label"] = [""]
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Molecule_Symm_label"] = (f'{col_n["B_Assignment_Sites_Molecule_Symm_label"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Atom_Labels"] = [""]
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Atom_Labels"] = (f'{col_n["B_Assignment_Sites_Atom_Labels"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Atom_Comment"] = ""
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_1_Atom_Comment"] = (f'{col_n["B_Assignment_Sites_Atom_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        else:
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Species_qty"] = len(selected_items)
                            for index_lvl_3 in range(1, len(selected_items) + 1):
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Specie_{index_lvl_3}_UID"] = attribution(sheet, f'{col_n["B_Assignment_Transition_Specie_UID"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Transition_Specie_{index_lvl_3}_UID"] = (f'{col_n["B_Assignment_Transition_Specie_UID"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Molecule_labels"] = multiple_collect(sheet, f'{col_n["B_Assignment_Sites_Molecule_labels"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1], True)
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Molecule_labels"] = multiple_position(sheet, f'{col_n["B_Assignment_Sites_Molecule_labels"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1], True)
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Molecule_Symm_label"] = multiple_collect(sheet, f'{col_n["B_Assignment_Sites_Molecule_Symm_label"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1], True)
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Molecule_Symm_label"] = multiple_position(sheet, f'{col_n["B_Assignment_Sites_Molecule_Symm_label"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1], True)
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Atom_Labels"] = multiple_collect(sheet, f'{col_n["B_Assignment_Sites_Atom_Labels"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1])
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Atom_Labels"] = multiple_position(sheet, f'{col_n["B_Assignment_Sites_Atom_Labels"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1])
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Atom_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Sites_Atom_Comment"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Site_{index_lvl_3}_Atom_Comment"] = (f'{col_n["B_Assignment_Sites_Atom_Comment"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                        # Transition : electronic_modes
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Types"] = multiple_collect(sheet, f"{col_n['B_Assignment_Electronic_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], empty_lines=True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Types"] = multiple_position(sheet, f"{col_n['B_Assignment_Electronic_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Labels"] = multiple_collect(sheet, f"{col_n['B_Assignment_Electronic_Labels']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], empty_lines=True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Labels"] = multiple_position(sheet, f"{col_n['B_Assignment_Electronic_Labels']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Electronic_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Electronic_Comment"] = (f'{col_n["B_Assignment_Electronic_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Transition : vibration_modes
                        selected_items = sub_multiples(sheet, [f"{col_n['B_Assignment_Vibration_Types']}", f"{col_n['B_Assignment_Vibration_Label']}"], assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        if len(selected_items) == 0:
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibrations_qty"] = 1
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Types"] = ""
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Types"] = (f'{col_n["B_Assignment_Vibration_Types"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Label"] = ""
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Label"] = (f'{col_n["B_Assignment_Vibration_Label"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Bonds"] = [""]
                            position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_1_Bonds"] = (f'{col_n["B_Assignment_Vibration_Bonds"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        else:
                            XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibrations_qty"] = len(selected_items)
                            for index_lvl_3 in range(1, len(selected_items) + 1):
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Types"] = attribution(sheet, f'{col_n["B_Assignment_Vibration_Types"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Types"] = (f'{col_n["B_Assignment_Vibration_Types"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Label"] = attribution(sheet, f'{col_n["B_Assignment_Vibration_Label"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Label"] = (f'{col_n["B_Assignment_Vibration_Label"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Bonds"] = multiple_collect(sheet, f"{col_n['B_Assignment_Vibration_Bonds']}", selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1])
                                position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibration_{index_lvl_3}_Bonds"] = multiple_position(sheet, f"{col_n['B_Assignment_Vibration_Bonds']}", selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibrations_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Vibration_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Vibrations_Comment"] = (f'{col_n["B_Assignment_Vibration_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Transition : rotation_modes
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Types"] = multiple_collect(sheet, f"{col_n['B_Assignment_Rotation_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Types"] = multiple_position(sheet, f"{col_n['B_Assignment_Rotation_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Label"] = multiple_collect(sheet, f"{col_n['B_Assignment_Rotation_Label']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Label"] = multiple_position(sheet, f"{col_n['B_Assignment_Rotation_Label']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Rotation_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Rotation_Comment"] = (f'{col_n["B_Assignment_Rotation_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Transition : phonon_modes
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Types"] = multiple_collect(sheet, f"{col_n['B_Assignment_Phonon_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Types"] = multiple_position(sheet, f"{col_n['B_Assignment_Phonon_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Label"] = multiple_collect(sheet, f"{col_n['B_Assignment_Phonon_Label']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Label"] = multiple_position(sheet, f"{col_n['B_Assignment_Phonon_Label']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Comment"] = attribution(sheet, f'{col_n["B_Assignment_Phonon_Comment"]}{assignment_list[index_lvl_2 - 1][1]}')
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Phonon_Comment"] = (f'{col_n["B_Assignment_Phonon_Comment"]}', assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][1])
                        # Transition : resonances
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Types"] = multiple_collect(sheet, f"{col_n['B_Assignment_Resonances_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Types"] = (f"{col_n['B_Assignment_Resonances_Types']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Band"] = multiple_collect(sheet, f"{col_n['B_Assignment_Resonances_Band']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Band"] = (f"{col_n['B_Assignment_Resonances_Band']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Nb"] = multiple_collect(sheet, f"{col_n['B_Assignment_Resonances_Nb']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Nb"] = (f"{col_n['B_Assignment_Resonances_Nb']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                        XLSX_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Comment"] = multiple_collect(sheet, f"{col_n['B_Assignment_Resonances_Comment']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2], True)
                        position_data[f"B_{index_lvl_1}_Assignment_{index_lvl_2}_Resonances_Comment"] = (f"{col_n['B_Assignment_Resonances_Comment']}", assignment_list[index_lvl_2 - 1][1], assignment_list[index_lvl_2 - 1][2])
                # Bands : Publications
                XLSX_data[f"B_{index_lvl_1}_Publications_Comments"] = attribution(sheet, f'{col_n["B_Publications_Comments"]}{band_list[index_lvl_1 - 1][1]}')
                position_data[f"B_{index_lvl_1}_Publications_Comments"] = (f'{col_n["B_Publications_Comments"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                temp_result = multiple_ordered_collect(sheet, f'{col_n["B_Publications_UID"]}', f'{col_n["B_Publications_Nb"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if type(temp_result) == str:
                    raise Exception(temp_result)
                XLSX_data[f"B_{index_lvl_1}_Publications_UID"] = temp_result
                position_data[f"B_{index_lvl_1}_Publications_UID"] = multiple_position(sheet, f'{col_n["B_Publications_UID"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if XLSX_data[f"B_{index_lvl_1}_Publications_UID"] == [""]:
                    null_flag = False
                    non_empty_flag = 0
                    for in_dex in range(band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2] + 1):
                        if attribution(sheet, f'{col_n["B_Publications_UID"]}{in_dex}'):
                            non_empty_flag = non_empty_flag + 1
                            if attribution(sheet, f'{col_n["B_Publications_UID"]}{in_dex}') == "NULL":
                                null_flag = True
                    if non_empty_flag == 1 and null_flag:
                        XLSX_data[f"B_{index_lvl_1}_Publications_UID"] = ["NULL"]
                # Bands : data SSHADE (UID)
                temp_result = multiple_ordered_collect(sheet, f'{col_n["B_Publications_SSHADE_UID"]}', f'{col_n["B_Publications_SSHADE_Nb"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if type(temp_result) == str:
                    raise Exception(temp_result)
                XLSX_data[f"B_{index_lvl_1}_Publications_SSHADE_UID"] = temp_result
                position_data[f"B_{index_lvl_1}_Publications_SSHADE_UID"] = multiple_position(sheet, f'{col_n["B_Publications_SSHADE_UID"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                # Bands : data URL (UID)
                temp_result = multiple_ordered_collect(sheet, f'{col_n["B_Publications_Data_URL"]}', f'{col_n["B_Publications_Data_Nb"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if type(temp_result) == str:
                    raise Exception(temp_result)
                XLSX_data[f"B_{index_lvl_1}_Publications_Data_URL"] = temp_result
                position_data[f"B_{index_lvl_1}_Publications_Data_URL"] = multiple_position(sheet, f'{col_n["B_Publications_Data_URL"]}', band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                # Characteristics
                temp_result = sub_multiples_ordered(sheet, [f"{col_n['B_Characteristics_T_Value']}", f"{col_n['B_Characteristics_Nb']}"], f"{col_n['B_Characteristics_Nb']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][2])
                if type(temp_result) == str:
                    raise Exception(temp_result)
                chars_list = temp_result
                if len(chars_list) == 0:
                    XLSX_data[f"B_{index_lvl_1}_Characteristics_qty"] = 1
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Nb"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Nb"] = (f"{col_n['B_Characteristics_Nb']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Composition"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Composition"] = (f"{col_n['B_Characteristics_Composition']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Texture"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Texture"] = (f"{col_n['B_Characteristics_Texture']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : temperature
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Unit"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Unit"] = (f"{col_n['B_Characteristics_T_Unit']}", start_line + 3, start_line + 3)
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Value"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Value"] = (f"{col_n['B_Characteristics_T_Value']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Error"] = (f"{col_n['B_Characteristics_T_Error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Formation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Formation"] = (f"{col_n['B_Characteristics_T_Formation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Max"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Max"] = (f"{col_n['B_Characteristics_T_Max']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_T_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_T_Comment"] = (f"{col_n['B_Characteristics_T_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : pressure
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Unit"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Unit"] = (f"{col_n['B_Characteristics_P_Unit']}", start_line + 3, start_line + 3)
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Value"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Value"] = (f"{col_n['B_Characteristics_P_Value']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Error"] = (f"{col_n['B_Characteristics_P_Error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Formation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Formation"] = (f"{col_n['B_Characteristics_P_Formation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Max"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Max"] = (f"{col_n['B_Characteristics_P_Max']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Stress_type"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Stress_type"] = (f"{col_n['B_Characteristics_P_Stress_type']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_P_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_P_Comment"] = (f"{col_n['B_Characteristics_P_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : excitation
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Laser_excitation_Wavelength_Unit"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Laser_excitation_Wavelength_Unit"] = (f"{col_n['B_Characteristics_Laser_excitation_Wavelength_Unit']}", start_line + 2, start_line + 2)
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Laser_excitation_Wavelength"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Laser_excitation_Wavelength"] = (f"{col_n['B_Characteristics_Laser_excitation_Wavelength']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Sample_Orient_mode"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Sample_Orient_mode"] = (f"{col_n['B_Characteristics_Sample_Orient_mode']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Sample_Orient"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Sample_Orient"] = (f"{col_n['B_Characteristics_Sample_Orient']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Polarization_Orient_mode"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Polarization_Orient_mode"] = (f"{col_n['B_Characteristics_Polarization_Orient_mode']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Polarization_Orient"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Polarization_Orient"] = (f"{col_n['B_Characteristics_Polarization_Orient']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Excitation_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Excitation_Comment"] = (f"{col_n['B_Characteristics_Excitation_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : method
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Methods_qty"] = 1
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Types"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Types"] = (f"{col_n['B_Characteristics_Method_Types']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Description"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Description"] = (f"{col_n['B_Characteristics_Method_Description']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Fit_Fct_type"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Fit_Fct_type"] = (f"{col_n['B_Characteristics_Method_Fit_Fct_type']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Fit_parameters"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Method_1_Fit_parameters"] = (f"{col_n['B_Characteristics_Method_Fit_parameters']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : overlap
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Methods_Overlap"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Methods_Overlap"] = (f"{col_n['B_Characteristics_Methods_Overlap']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : position
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak_method"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak_method"] = (f"{col_n['B_Characteristics_Position_Peak_method']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak"] = (f"{col_n['B_Characteristics_Position_Peak']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Peak_error"] = (f"{col_n['B_Characteristics_Position_Peak_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center_method"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center_method"] = (f"{col_n['B_Characteristics_Position_Center_method']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center"] = (f"{col_n['B_Characteristics_Position_Center']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Center_error"] = (f"{col_n['B_Characteristics_Position_Center_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Evaluation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Evaluation"] = (f"{col_n['B_Characteristics_Position_Evaluation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Position_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Position_Comment"] = (f"{col_n['B_Characteristics_Position_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : width
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Method"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Method"] = (f"{col_n['B_Characteristics_Width_Method']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_FWHM"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_FWHM"] = (f"{col_n['B_Characteristics_Width_FWHM']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_FWHM_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_FWHM_error"] = (f"{col_n['B_Characteristics_Width_FWHM_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Shape"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Shape"] = (f"{col_n['B_Characteristics_Width_Shape']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Asymm_factor"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Asymm_factor"] = (f"{col_n['B_Characteristics_Width_Asymm_factor']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Asymm_factor_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Asymm_factor_error"] = (f"{col_n['B_Characteristics_Width_Asymm_factor_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Evaluation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Evaluation"] = (f"{col_n['B_Characteristics_Width_Evaluation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Width_Comments"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Width_Comments"] = (f"{col_n['B_Characteristics_Width_Comments']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : peak_intensity
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Method"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Method"] = (f"{col_n['B_Characteristics_Peak_intensity_Method']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_sp"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_sp"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_sp_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Abs_coef_sp_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Relative"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Relative"] = (f"{col_n['B_Characteristics_Peak_intensity_Relative']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Relative_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Relative_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Relative_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Strength"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Strength"] = (f"{col_n['B_Characteristics_Peak_intensity_Strength']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Evaluation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Evaluation"] = (f"{col_n['B_Characteristics_Peak_intensity_Evaluation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Peak_intensity_Comment"] = (f"{col_n['B_Characteristics_Peak_intensity_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : integrated_intensity
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Method"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Method"] = (f"{col_n['B_Characteristics_Integrated_intensity_Method']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_sp"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_sp"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_sp_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Abs_coef_sp_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Relative"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Relative"] = (f"{col_n['B_Characteristics_Integrated_intensity_Relative']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Relative_error"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Relative_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Relative_error']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Strength"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Strength"] = (f"{col_n['B_Characteristics_Integrated_intensity_Strength']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Evaluation"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Evaluation"] = (f"{col_n['B_Characteristics_Integrated_intensity_Evaluation']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Comment"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Integrated_intensity_Comment"] = (f"{col_n['B_Characteristics_Integrated_intensity_Comment']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                    # Characteristics : bandlist_nominal_flag
                    XLSX_data[f"B_{index_lvl_1}_Characteristic_1_Bandlist_flag"] = ""
                    position_data[f"B_{index_lvl_1}_Characteristic_1_Bandlist_flag"] = (f"{col_n['B_Characteristics_Bandlist_flag']}", band_list[index_lvl_1 - 1][1], band_list[index_lvl_1 - 1][1])
                else:
                    XLSX_data[f"B_{index_lvl_1}_Characteristics_qty"] = len(chars_list)
                    for index_lvl_2 in range(1, len(chars_list) + 1):
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Nb"] = attribution(sheet, f"{col_n['B_Characteristics_Nb']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Nb"] = (f"{col_n['B_Characteristics_Nb']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Composition"] = attribution(sheet, f"{col_n['B_Characteristics_Composition']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Composition"] = (f"{col_n['B_Characteristics_Composition']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Texture"] = attribution(sheet, f"{col_n['B_Characteristics_Texture']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Texture"] = (f"{col_n['B_Characteristics_Texture']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : temperature
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Unit"] = attribution(sheet, f"{col_n['B_Characteristics_T_Unit']}{start_line + 3}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Unit"] = (f"{col_n['B_Characteristics_T_Unit']}", start_line + 3, start_line + 3)
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Value"] = attribution(sheet, f"{col_n['B_Characteristics_T_Value']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Value"] = (f"{col_n['B_Characteristics_T_Value']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Error"] = attribution(sheet, f"{col_n['B_Characteristics_T_Error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Error"] = (f"{col_n['B_Characteristics_T_Error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Formation"] = attribution(sheet, f"{col_n['B_Characteristics_T_Formation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Formation"] = (f"{col_n['B_Characteristics_T_Formation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Max"] = attribution(sheet, f"{col_n['B_Characteristics_T_Max']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Max"] = (f"{col_n['B_Characteristics_T_Max']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_T_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_T_Comment"] = (f"{col_n['B_Characteristics_T_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : pressure
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Unit"] = attribution(sheet, f"{col_n['B_Characteristics_P_Unit']}{start_line + 3}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Unit"] = (f"{col_n['B_Characteristics_P_Unit']}", start_line + 3, start_line + 3)
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Value"] = attribution(sheet, f"{col_n['B_Characteristics_P_Value']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Value"] = (f"{col_n['B_Characteristics_P_Value']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Error"] = attribution(sheet, f"{col_n['B_Characteristics_P_Error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Error"] = (f"{col_n['B_Characteristics_P_Error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Formation"] = attribution(sheet, f"{col_n['B_Characteristics_P_Formation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Formation"] = (f"{col_n['B_Characteristics_P_Formation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Max"] = attribution(sheet, f"{col_n['B_Characteristics_P_Max']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Max"] = (f"{col_n['B_Characteristics_P_Max']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Stress_type"] = attribution(sheet, f"{col_n['B_Characteristics_P_Stress_type']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Stress_type"] = (f"{col_n['B_Characteristics_P_Stress_type']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_P_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_P_Comment"] = (f"{col_n['B_Characteristics_P_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : excitation
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Laser_excitation_Wavelength_Unit"] = attribution(sheet, f"{col_n['B_Characteristics_Laser_excitation_Wavelength_Unit']}{start_line + 2}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Laser_excitation_Wavelength_Unit"] = (f"{col_n['B_Characteristics_Laser_excitation_Wavelength_Unit']}", start_line + 2, start_line + 2)
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Laser_excitation_Wavelength"] = attribution(sheet, f"{col_n['B_Characteristics_Laser_excitation_Wavelength']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Laser_excitation_Wavelength"] = (f"{col_n['B_Characteristics_Laser_excitation_Wavelength']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Sample_Orient_mode"] = attribution(sheet, f"{col_n['B_Characteristics_Sample_Orient_mode']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Sample_Orient_mode"] = (f"{col_n['B_Characteristics_Sample_Orient_mode']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Sample_Orient"] = attribution(sheet, f"{col_n['B_Characteristics_Sample_Orient']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Sample_Orient"] = (f"{col_n['B_Characteristics_Sample_Orient']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Polarization_Orient_mode"] = attribution(sheet, f"{col_n['B_Characteristics_Polarization_Orient_mode']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Polarization_Orient_mode"] = (f"{col_n['B_Characteristics_Polarization_Orient_mode']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Polarization_Orient"] = attribution(sheet, f"{col_n['B_Characteristics_Polarization_Orient']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Polarization_Orient"] = (f"{col_n['B_Characteristics_Polarization_Orient']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Excitation_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_Excitation_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Excitation_Comment"] = (f"{col_n['B_Characteristics_Excitation_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : method
                        selected_items = sub_multiples(sheet, [f"{col_n['B_Characteristics_Method_Types']}", f"{col_n['B_Characteristics_Method_Description']}", f"{col_n['B_Characteristics_Method_Fit_Fct_type']}", f"{col_n['B_Characteristics_Method_Fit_parameters']}"], chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][2])
                        if len(selected_items) == 0:
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Methods_qty"] = 1
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Types"] = ""
                            position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Types"] = (f"{col_n['B_Characteristics_Method_Types']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Description"] = ""
                            position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Description"] = (f"{col_n['B_Characteristics_Method_Description']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Fit_Fct_type"] = ""
                            position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Fit_Fct_type"] = (f"{col_n['B_Characteristics_Method_Fit_Fct_type']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Fit_parameters"] = ""
                            position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_1_Fit_parameters"] = (f"{col_n['B_Characteristics_Method_Fit_parameters']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        else:
                            XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Methods_qty"] = len(selected_items)
                            for index_lvl_3 in range(1, len(selected_items) + 1):
                                XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Types"] = attribution(sheet, f'{col_n["B_Characteristics_Method_Types"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Types"] = (f'{col_n["B_Characteristics_Method_Types"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Description"] = attribution(sheet, f'{col_n["B_Characteristics_Method_Description"]}{selected_items[index_lvl_3 - 1][0]}')
                                position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Description"] = (f'{col_n["B_Characteristics_Method_Description"]}', selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Fit_Fct_type"] = attribution(sheet, f"{col_n['B_Characteristics_Method_Fit_Fct_type']}{selected_items[index_lvl_3 - 1][0]}")
                                position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Fit_Fct_type"] = (f"{col_n['B_Characteristics_Method_Fit_Fct_type']}", selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                                XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Fit_parameters"] = attribution(sheet, f"{col_n['B_Characteristics_Method_Fit_parameters']}{selected_items[index_lvl_3 - 1][0]}")
                                position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Method_{index_lvl_3}_Fit_parameters"] = (f"{col_n['B_Characteristics_Method_Fit_parameters']}", selected_items[index_lvl_3 - 1][0], selected_items[index_lvl_3 - 1][0])
                        # Characteristics : overlap
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Methods_Overlap"] = attribution(sheet, f"{col_n['B_Characteristics_Methods_Overlap']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Methods_Overlap"] = (f"{col_n['B_Characteristics_Methods_Overlap']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : position
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak_method"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Peak_method']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak_method"] = (f"{col_n['B_Characteristics_Position_Peak_method']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Peak']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak"] = (f"{col_n['B_Characteristics_Position_Peak']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak_error"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Peak_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Peak_error"] = (f"{col_n['B_Characteristics_Position_Peak_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center_method"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Center_method']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center_method"] = (f"{col_n['B_Characteristics_Position_Center_method']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Center']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center"] = (f"{col_n['B_Characteristics_Position_Center']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center_error"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Center_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Center_error"] = (f"{col_n['B_Characteristics_Position_Center_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Evaluation"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Evaluation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Evaluation"] = (f"{col_n['B_Characteristics_Position_Evaluation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_Position_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Position_Comment"] = (f"{col_n['B_Characteristics_Position_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : width
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Method"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Method']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Method"] = (f"{col_n['B_Characteristics_Width_Method']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_FWHM"] = attribution(sheet, f"{col_n['B_Characteristics_Width_FWHM']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_FWHM"] = (f"{col_n['B_Characteristics_Width_FWHM']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_FWHM_error"] = attribution(sheet, f"{col_n['B_Characteristics_Width_FWHM_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_FWHM_error"] = (f"{col_n['B_Characteristics_Width_FWHM_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Shape"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Shape']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Shape"] = (f"{col_n['B_Characteristics_Width_Shape']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Asymm_factor"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Asymm_factor']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Asymm_factor"] = (f"{col_n['B_Characteristics_Width_Asymm_factor']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Asymm_factor_error"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Asymm_factor_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Asymm_factor_error"] = (f"{col_n['B_Characteristics_Width_Asymm_factor_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Evaluation"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Evaluation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Evaluation"] = (f"{col_n['B_Characteristics_Width_Evaluation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Comments"] = attribution(sheet, f"{col_n['B_Characteristics_Width_Comments']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Width_Comments"] = (f"{col_n['B_Characteristics_Width_Comments']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : peak_intensity
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Method"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Method']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Method"] = (f"{col_n['B_Characteristics_Peak_intensity_Method']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Abs_coef']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_error"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_sp"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_sp"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_sp_error"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Abs_coef_sp_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Abs_coef_sp_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Relative"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Relative']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Relative"] = (f"{col_n['B_Characteristics_Peak_intensity_Relative']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Relative_error"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Relative_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Relative_error"] = (f"{col_n['B_Characteristics_Peak_intensity_Relative_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Strength"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Strength']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Strength"] = (f"{col_n['B_Characteristics_Peak_intensity_Strength']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Evaluation"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Evaluation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Evaluation"] = (f"{col_n['B_Characteristics_Peak_intensity_Evaluation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_Peak_intensity_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Peak_intensity_Comment"] = (f"{col_n['B_Characteristics_Peak_intensity_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : integrated_intensity
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Method"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Method']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Method"] = (f"{col_n['B_Characteristics_Integrated_intensity_Method']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_error"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_sp"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_sp"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_sp_error"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Abs_coef_sp_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Abs_coef_sp_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Relative"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Relative']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Relative"] = (f"{col_n['B_Characteristics_Integrated_intensity_Relative']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Relative_error"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Relative_error']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Relative_error"] = (f"{col_n['B_Characteristics_Integrated_intensity_Relative_error']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Strength"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Strength']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Strength"] = (f"{col_n['B_Characteristics_Integrated_intensity_Strength']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Evaluation"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Evaluation']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Evaluation"] = (f"{col_n['B_Characteristics_Integrated_intensity_Evaluation']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Comment"] = attribution(sheet, f"{col_n['B_Characteristics_Integrated_intensity_Comment']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Integrated_intensity_Comment"] = (f"{col_n['B_Characteristics_Integrated_intensity_Comment']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
                        # Characteristics : bandlist_nominal_flag
                        XLSX_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Bandlist_flag"] = attribution(sheet, f"{col_n['B_Characteristics_Bandlist_flag']}{chars_list[index_lvl_2 - 1][1]}")
                        position_data[f"B_{index_lvl_1}_Characteristic_{index_lvl_2}_Bandlist_flag"] = (f"{col_n['B_Characteristics_Bandlist_flag']}", chars_list[index_lvl_2 - 1][1], chars_list[index_lvl_2 - 1][1])
    return XLSX_data, position_data


# XML FILLING
def XML_filler(xlsx_workbook, bandlist_type):
    # VARS
    XLSX_data = XLSX_reader(xlsx_workbook, bandlist_type)[0]

    # INNER FUNCTIONS
    def accent_letters_replace(string_var):
        # special chars
        string_var = string_var.replace(">", "")
        string_var = string_var.replace("<", "")
        string_var = string_var.replace("'", "")
        string_var = string_var.replace('"', "")
        string_var = string_var.replace('&', "")
        string_var = string_var.replace(']]', "")
        return string_var

    # XML actions
    def fill_child_action(child_element, its_value):
        child_element.clear()
        if type(its_value) == str:
            if accent_letters_replace(its_value) == its_value:
                child_element.text = its_value
            else:
                child_element.text = etree.CDATA(its_value)
        else:
            child_element.text = str(its_value)

    def fill_multiple_action(parent, multiple_array, tag_name):
        if len(multiple_array) != 0:
            for child in parent:
                parent.remove(child)
            all_empty = True
            for doc_index in range(0, len(multiple_array)):
                if multiple_array[doc_index]:
                    all_empty = False
                    child_element = etree.SubElement(parent, tag_name)
                    fill_child_action(child_element, multiple_array[doc_index])
            if all_empty:
                child_element = etree.SubElement(parent, tag_name)
                fill_child_action(child_element, multiple_array[0])

    def fill_multiple_parent_action(parent, sub_parent_tag, children_tag_name_list, multiple_arraies_list, empty_tags=True):
        for child in parent:
            parent.remove(child)
        at_least_one = False
        for inner_doc_index in range(0, len(multiple_arraies_list[0])):
            if empty_tags:
                at_least_one = True
                child_element = etree.SubElement(parent, sub_parent_tag)
                for doc_index in range(0, len(multiple_arraies_list)):
                    sub_child_element = etree.SubElement(child_element, children_tag_name_list[doc_index])
                    fill_child_action(sub_child_element, multiple_arraies_list[doc_index][inner_doc_index])
            else:
                only_empty = True
                for doc_index in range(0, len(multiple_arraies_list)):
                    if multiple_arraies_list[doc_index][inner_doc_index] != "":
                        only_empty = False
                if not only_empty:
                    child_element = etree.SubElement(parent, sub_parent_tag)
                    for doc_index in range(0, len(multiple_arraies_list)):
                        at_least_one = True
                        sub_child_element = etree.SubElement(child_element, children_tag_name_list[doc_index])
                        fill_child_action(sub_child_element, multiple_arraies_list[doc_index][inner_doc_index])
        if not at_least_one:
            child_element = etree.SubElement(parent, sub_parent_tag)
            for doc_index in range(0, len(multiple_arraies_list)):
                sub_child_element = etree.SubElement(child_element, children_tag_name_list[doc_index])
                fill_child_action(sub_child_element, multiple_arraies_list[doc_index][0])
    # xml_template PARSE and FILL
    parser = etree.XMLParser(remove_blank_text=True)
    xml_root = etree.fromstring(xml_template.encode("utf8"), parser)
    uid = ""
    for child in xml_root.find("{http://sshade.eu/schema/import}bandlist").getchildren():
        # Bandlist
        # Bandlist: import_mode
        if child.tag == "{http://sshade.eu/schema/import}import_mode":
            fill_child_action(child, XLSX_data['BL_Import_mode'])
        # Bandlist: uid
        if child.tag == "{http://sshade.eu/schema/import}uid":
            if XLSX_data['BL_UID'].find("BANDLIST_") == -1:
                uid = "BANDLIST_"
            if XLSX_data['BL_UID'].find("ABS_") == -1 and XLSX_data['BL_UID'].find("RAMAN_") == -1:
                if bandlist_type == "ABS":
                    XLSX_data['BL_UID'] = "ABS_" + XLSX_data['BL_UID']
                elif bandlist_type == "RAMAN":
                    XLSX_data['BL_UID'] = "RAMAN_" + XLSX_data['BL_UID']
            uid = uid + XLSX_data['BL_UID']
            fill_child_action(child, uid)
        # Bandlist: type
        if child.tag == "{http://sshade.eu/schema/import}type":
            fill_child_action(child, XLSX_data['BL_Type'])
        # Bandlist: title
        if child.tag == "{http://sshade.eu/schema/import}title":
            fill_child_action(child, XLSX_data['BL_Title'])
        # Bandlist: description
        if child.tag == "{http://sshade.eu/schema/import}description":
            fill_child_action(child, XLSX_data['BL_Description'])
        # Bandlist: parameters_spectral
        if child.tag == "{http://sshade.eu/schema/import}parameters_spectral":
            for subchild_lvl_1 in child:
                # Bandlist: parameters_spectral unit
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}unit":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Spectral_Unit'])
                # Bandlist: parameters_spectral standard
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}standard":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Spectral_Standard'])
                # Bandlist: parameters_spectral range_types
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}range_types":
                    fill_multiple_action(subchild_lvl_1, XLSX_data['BL_Spectral_Range_types'], "type")
                # Bandlist: parameters_spectral ranges
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}ranges":
                    fill_multiple_parent_action(subchild_lvl_1, "range", ["min", "max"], [XLSX_data['BL_Spectral_Range_min'], XLSX_data['BL_Spectral_Range_max']], False)
                # Bandlist: parameters_spectral ranges
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}comments":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Spectral_Comments'])
        # Bandlist: reference_position
        if child.tag == "{http://sshade.eu/schema/import}reference_position":
            for subchild_lvl_1 in child:
                # Bandlist: reference_position electronic
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}electronic":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Spectral_Ref_pos_electronic'])
                # Bandlist: reference_position infrared
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}infrared":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Spectral_Ref_pos_absorption'])
        # Bandlist: constituent
        if child.tag == "{http://sshade.eu/schema/import}constituent":
            for subchild_lvl_1 in child:
                # Bandlist: constituent uid
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}uid":
                    fill_child_action(subchild_lvl_1, "CONST_" + XLSX_data['BL_Constituent_UID'])
                # Bandlist: constituent primary_specie_uid
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}primary_specie_uid":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Constituent_Primary_specie_UID'])
                # Bandlist: constituent comments
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}comments":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Constituent_Comments'])
        # Bandlist: previous_version
        if child.tag == "{http://sshade.eu/schema/import}previous_version":
            for subchild_lvl_1 in child:
                # Bandlist: previous_version status
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}status":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Versions_Previous_version_status'])
                # Bandlist: previous_version comments
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}comments":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Versions_Comments'])
        # Bandlist: history
        if child.tag == "{http://sshade.eu/schema/import}history":
            fill_child_action(child, XLSX_data['BL_Versions_Current_version_history'])
        # Bandlist: parent_experiments
        if child.tag == "{http://sshade.eu/schema/import}parent_experiments":
            for i_ndex, value in enumerate(XLSX_data['BL_Parents_Exp_UID']):
                if value and value != "NULL" and value.find("EXPERIMENT_") == -1:
                    XLSX_data['BL_Parents_Exp_UID'][i_ndex] = "EXPERIMENT_" + value
            fill_multiple_action(child, XLSX_data['BL_Parents_Exp_UID'], "uid")
        # Bandlist: parent_spectra
        if child.tag == "{http://sshade.eu/schema/import}parent_spectra":
            for i_ndex, value in enumerate(XLSX_data['BL_Parents_Spectra_UID']):
                if value and value != "NULL" and value.find("SPECTRUM_") == -1:
                    XLSX_data['BL_Parents_Spectra_UID'][i_ndex] = "SPECTRUM_" + value
            fill_multiple_action(child, XLSX_data['BL_Parents_Spectra_UID'], "uid")
        # Bandlist: parent_spectra_comments
        if child.tag == "{http://sshade.eu/schema/import}parent_spectra_comments":
            fill_child_action(child, XLSX_data['BL_Parents_Comments'])
        # Bandlist: analysis
        if child.tag == "{http://sshade.eu/schema/import}analysis":
            fill_child_action(child, XLSX_data['BL_Analysis'])
        # Bandlist: comments (globals)
        if child.tag == "{http://sshade.eu/schema/import}comments":
            fill_child_action(child, XLSX_data['BL_Global_comments'])
        # Bandlist: quality_flag
        if child.tag == "{http://sshade.eu/schema/import}quality_flag":
            fill_child_action(child, XLSX_data['BL_Validation_Quality'])
        # Bandlist: date_validated
        if child.tag == "{http://sshade.eu/schema/import}date_validated":
            fill_child_action(child, XLSX_data['BL_Validation_Date_validated'])
        # Bandlist: validators
        if child.tag == "{http://sshade.eu/schema/import}validators":
            if is_list_with_only_empty_strings(XLSX_data['BL_Validation_Validators_UID']):
                fill_multiple_action(child, [""], "experimentalist_uid")
            else:
                fill_multiple_action(child, ["EXPER_" + XLSX_data['BL_Validation_Validators_UID'][doc_index] for doc_index in range(0, len(XLSX_data['BL_Validation_Validators_UID']))], "experimentalist_uid")
        # Bandlist: documentations
        if child.tag == "{http://sshade.eu/schema/import}documentations":
            if XLSX_data['BL_Documentation_names'] or XLSX_data['BL_Documentation_files']:
                fill_multiple_parent_action(child, "documentation", ["name", "filename"], [XLSX_data['BL_Documentation_names'], XLSX_data['BL_Documentation_files']], False)
        # Bandlist: original_data_filename
        if child.tag == "{http://sshade.eu/schema/import}original_data_filename":
            fill_child_action(child, XLSX_data['BL_Original_data_filename'])
        # Bandlist: export_filename
        if child.tag == "{http://sshade.eu/schema/import}export_filename":
            fill_child_action(child, XLSX_data['BL_Export_filename'])
        # Bandlist: preview
        if child.tag == "{http://sshade.eu/schema/import}preview":
            for subchild_lvl_1 in child:
                # Bandlist: preview x
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}x":
                    fill_child_action(subchild_lvl_1, "")
                    subchild_lvl_1.set("axis", XLSX_data['BL_Preview_x_Axis'])
                    subchild_lvl_1.set("unit", XLSX_data['BL_Preview_x_Unit'])
                    subchild_lvl_1.set("min", XLSX_data['BL_Preview_x_Min'])
                    subchild_lvl_1.set("max", XLSX_data['BL_Preview_x_Max'])
                # Bandlist: preview y
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}y":
                    fill_child_action(subchild_lvl_1, "")
                    subchild_lvl_1.set("axis", XLSX_data['BL_Preview_y_Axis'])
                    subchild_lvl_1.set("unit", XLSX_data['BL_Preview_y_Unit'])
                    subchild_lvl_1.set("min", XLSX_data['BL_Preview_y_Min'])
                    subchild_lvl_1.set("max", XLSX_data['BL_Preview_y_Max'])
                # Bandlist: preview y_rel
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}yrel":
                    fill_child_action(subchild_lvl_1, "")
                    subchild_lvl_1.set("axis", XLSX_data['BL_Preview_y_rel_Axis'])
                    subchild_lvl_1.set("min", XLSX_data['BL_Preview_y_rel_Min'])
                    subchild_lvl_1.set("max", XLSX_data['BL_Preview_y_rel_Max'])
                # Bandlist: preview type
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}type":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Preview_Type'])
                # Bandlist: preview filename
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}filename":
                    fill_child_action(subchild_lvl_1, XLSX_data['BL_Preview_Filename'])
        # Bandlist: structure
        if child.tag == "{http://sshade.eu/schema/import}structure":
            for subchild_lvl_1 in child:
                # Bandlist: structure sections
                if subchild_lvl_1.tag == "{http://sshade.eu/schema/import}sections":
                    subchild_lvl_1.set("variable_parameter", XLSX_data['BL_Sections_Var_param'])
                    for child_element in subchild_lvl_1:
                        subchild_lvl_1.remove(child_element)
                    for section in range(1, XLSX_data['BL_Sections_qty'] + 1):
                        subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "section")
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "title")
                        fill_child_action(subchild_lvl_3, XLSX_data[f'BL_Section_{section}_Title'])
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "description")
                        fill_child_action(subchild_lvl_3, XLSX_data[f'BL_Section_{section}_Description'])
                        if is_list_with_only_empty_strings(XLSX_data[f'BL_Section_{section}_Bands_UID']):
                            subchild_lvl_3 = etree.Comment(' SUBSECTIONS ')
                            subchild_lvl_2.insert(3, subchild_lvl_3)
                            subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "subsections")
                            subchild_lvl_3.set("variable_parameter", XLSX_data[f'BL_Section_{section}_Var_param'])
                            if XLSX_data[f'BL_Section_{section}_Sub_sections_qty'] != 0:
                                for sub_section in range(1, XLSX_data[f'BL_Section_{section}_Sub_sections_qty'] + 1):
                                    subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "subsection")
                                    subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "title")
                                    fill_child_action(subchild_lvl_5, XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Title'])
                                    subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "description")
                                    fill_child_action(subchild_lvl_5, XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Description'])
                                    subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "bands")
                                    if is_list_with_only_empty_strings(XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID']):
                                        XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID'] = [""]
                                    else:
                                        XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID'] = ["BAND_" + XLSX_data["BL_UID"] + "_" + XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID'][i] for i in range(0, len(XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID']))]
                                    fill_multiple_action(subchild_lvl_5, XLSX_data[f'BL_Section_{section}_Sub_section_{sub_section}_Bands_UID'], "band_uid")
                            else:
                                subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "subsection")
                                subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "title")
                                fill_child_action(subchild_lvl_5, "")
                                subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "description")
                                fill_child_action(subchild_lvl_5, "")
                                subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "bands")
                                fill_multiple_action(subchild_lvl_5, [""], "band_uid")
                        else:
                            subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "bands")
                            XLSX_data[f'BL_Section_{section}_Bands_UID'] = ["BAND_" + XLSX_data["BL_UID"] + "_" + XLSX_data[f'BL_Section_{section}_Bands_UID'][i] for i in range(0, len(XLSX_data[f'BL_Section_{section}_Bands_UID']))]
                            fill_multiple_action(subchild_lvl_3, XLSX_data[f'BL_Section_{section}_Bands_UID'], "band_uid")
        # Bands
        if child.tag == "{http://sshade.eu/schema/import}bands":
            for inner_child in child:
                child.remove(inner_child)
            for band_number in range(1, XLSX_data['B_qty'] + 1):
                if XLSX_data[f'B_{band_number}_Import_mode'] not in ["ignore", "draft"]:
                    # bands band
                    subchild_lvl_1 = etree.SubElement(child, "band")
                    # bands band import_mode
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "import_mode")
                    fill_child_action(subchild_lvl_2, XLSX_data[f'B_{band_number}_Import_mode'])
                    # bands band uid
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "uid")
                    if XLSX_data["BL_UID"].find("ABS") != -1 or XLSX_data["BL_UID"].find("RAMAN") != -1:
                        fill_child_action(subchild_lvl_2, 'BAND_' + XLSX_data["BL_UID"] + "_" + XLSX_data[f"B_{band_number}_UID"])
                    else:
                        fill_child_action(subchild_lvl_2, 'BAND_' + bandlist_type + '_' + XLSX_data["BL_UID"] + "_" + XLSX_data[f"B_{band_number}_UID"])
                    # bands band comments
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "comments")
                    fill_child_action(subchild_lvl_2, XLSX_data[f'B_{band_number}_Comment'])
                    # bands band ASSIGNMENTS_comment
                    subchild_lvl_2 = etree.Comment(' TRANSITION ASSIGNMENTS AND MODES ')
                    subchild_lvl_1.insert(3, subchild_lvl_2)
                    # bands band assignments
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "assignments")
                    for assignment_number in range(1, XLSX_data[f"B_{band_number}_Assignments_qty"] + 1):
                        # bands band assignments assignment
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "assignment")
                        # bands band assignments assignment number
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "number")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Number'])
                        # bands band assignments assignment label
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "label")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Label'])
                        # bands band assignments assignment symmetry_label
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "symmetry_label")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Symmetry'])
                        # bands band assignments assignment category
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "category")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Category'])
                        # bands band assignments assignment method
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "method")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Method'])
                        # bands band assignments assignment level
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "level")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Level'])
                        # bands band assignments assignment evaluation
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "evaluation")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Evaluation'])
                        # bands band assignments assignment comments
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "comments")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Comment'])
                        # bands band assignments assignment MULTIPLICITY_comment
                        subchild_lvl_4 = etree.Comment(' TRANSITION: MULTIPLICITY AND DEGENERACY ')
                        subchild_lvl_3.insert(7, subchild_lvl_4)
                        # bands band assignments assignment multiplicities
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "multiplicities")
                        if is_list_with_only_empty_strings(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types']):
                            # bands band assignments assignment multiplicity
                            subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "multiplicity")
                            # bands band assignments assignment multiplicity type
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "type")
                            fill_child_action(subchild_lvl_6, "")
                            # bands band assignments assignment multiplicity degeneracy
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "degeneracy")
                            fill_child_action(subchild_lvl_6, "")
                            # bands band assignments assignment multiplicity other_band_uid
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "other_band_uid")
                            fill_child_action(subchild_lvl_6, "")
                        else:
                            for multiplicity_number in range(0, len(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types'])):
                                if XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types'][multiplicity_number] != '' or XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Degeneracy'][multiplicity_number] != '' or XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number] != '':
                                    # bands band assignments assignment multiplicity
                                    subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "multiplicity")
                                    # bands band assignments assignment multiplicity type
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "type")
                                    fill_child_action(subchild_lvl_6, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types'][multiplicity_number])
                                    # bands band assignments assignment multiplicity degeneracy
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "degeneracy")
                                    fill_child_action(subchild_lvl_6, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Degeneracy'][multiplicity_number])
                                    # bands band assignments assignment multiplicity other_band_uid
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "other_band_uid")
                                    if XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number]:
                                        if XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number] == "NULL" or XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number].find("BAND_") != -1:
                                            fill_child_action(subchild_lvl_6, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number])
                                        else:
                                            fill_child_action(subchild_lvl_6, 'BAND_' + XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band'][multiplicity_number])
                                    else:
                                        fill_child_action(subchild_lvl_6, "")
                        # bands band assignments assignment contribution_level
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "contribution_level")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Contribution_Level'])
                        # bands band assignments assignment contribution_comments
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "contribution_comments")
                        fill_child_action(subchild_lvl_4, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Contribution_Comment'])
                        # bands band assignments assignment TRANSITION_comment
                        subchild_lvl_4 = etree.Comment(' TRANSITION MODES ')
                        subchild_lvl_3.insert(11, subchild_lvl_4)
                        # bands band assignments assignment transition
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "transition")
                        # bands band assignments assignment transition primary_species
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "primary_species")
                        for primary_specie_number in range(1, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Transition_Species_qty'] + 1):
                            # bands band assignments assignment transition primary_specie
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "primary_specie")
                            # bands band assignments assignment transition primary_specie uid
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "uid")
                            fill_child_action(subchild_lvl_7, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Transition_Specie_{primary_specie_number}_UID'])
                            # bands band assignments assignment transition primary_specie crystal_molecule_sites
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "crystal_molecule_sites")
                            if is_list_with_only_empty_strings(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_labels']) and is_list_with_only_empty_strings(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_Symm_label']):
                                # bands band assignments assignment transition primary_specie crystal_molecule_sites site
                                subchild_lvl_8 = etree.SubElement(subchild_lvl_7, "site")
                                # bands band assignments assignment transition primary_specie crystal_molecule_sites site label
                                subchild_lvl_9 = etree.SubElement(subchild_lvl_8, "label")
                                fill_child_action(subchild_lvl_9, "")
                                # bands band assignments assignment transition primary_specie crystal_molecule_sites site symmetry_label
                                subchild_lvl_9 = etree.SubElement(subchild_lvl_8, "symmetry_label")
                                fill_child_action(subchild_lvl_9, "")
                            else:
                                for site_number in range(0, len(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_labels'])):
                                    if XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_labels'][site_number] != "" or XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_Symm_label'][site_number]:
                                        # bands band assignments assignment transition primary_specie crystal_molecule_sites site
                                        subchild_lvl_8 = etree.SubElement(subchild_lvl_7, "site")
                                        # bands band assignments assignment transition primary_specie crystal_molecule_sites site label
                                        subchild_lvl_9 = etree.SubElement(subchild_lvl_8, "label")
                                        fill_child_action(subchild_lvl_9, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_labels'][site_number])
                                        # bands band assignments assignment transition primary_specie crystal_molecule_sites site symmetry_label
                                        subchild_lvl_9 = etree.SubElement(subchild_lvl_8, "symmetry_label")
                                        fill_child_action(subchild_lvl_9, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Molecule_Symm_label'][site_number])
                            # bands band assignments assignment transition primary_specie crystal_sites
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "crystal_sites")
                            for site_number in range(0, len(XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Atom_Labels'])):
                                # bands band assignments assignment transition primary_specie crystal_sites label
                                subchild_lvl_8 = etree.SubElement(subchild_lvl_7, "label")
                                fill_child_action(subchild_lvl_8, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Atom_Labels'][site_number])
                            # bands band assignments assignment transition primary_specie sites_comments
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "sites_comments")
                            fill_child_action(subchild_lvl_7, XLSX_data[f'B_{band_number}_Assignment_{assignment_number}_Site_{primary_specie_number}_Atom_Comment'])
                        # bands band assignments assignment transition ELECTRONIC_comment
                        subchild_lvl_5 = etree.Comment(' TRANSITION: ELECTRONIC ')
                        subchild_lvl_4.insert(1, subchild_lvl_5)
                        # bands band assignments assignment transition electronic_modes
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "electronic_modes")
                        if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types"]) and is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Labels"]):
                            # bands band assignments assignment transition electronic_modes mode
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                            # bands band assignments assignment transition electronic_modes mode label
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                            fill_child_action(subchild_lvl_7, "")
                            # bands band assignments assignment transition electronic_modes mode type
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                            fill_child_action(subchild_lvl_7, "")
                        else:
                            for mode_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types"])):
                                if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types"][mode_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Labels"][mode_number] != "":
                                    # bands band assignments assignment transition electronic_modes mode
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                                    # bands band assignments assignment transition electronic_modes mode label
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Labels"][mode_number])
                                    # bands band assignments assignment transition electronic_modes mode type
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types"][mode_number])
                        # bands band assignments assignment transition electronic_modes comments
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                        fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Comment"])
                        # bands band assignments assignment transition VIBRATION_comment
                        subchild_lvl_5 = etree.Comment(' TRANSITION: VIBRATION MODES ')
                        subchild_lvl_4.insert(3, subchild_lvl_5)
                        # bands band assignments assignment transition vibration_modes
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "vibration_modes")
                        for mode_number in range(1, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibrations_qty"] + 1):
                            # bands band assignments assignment transition vibration_modes mode
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                            # bands band assignments assignment transition vibration_modes mode label
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                            fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{mode_number}_Label"])
                            # bands band assignments assignment transition vibration_modes mode type
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                            fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{mode_number}_Types"])
                            # bands band assignments assignment transition vibration_modes mode chemical_bonds
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "chemical_bonds")
                            for bonds_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{mode_number}_Bonds"])):
                                # bands band assignments assignment transition vibration_modes mode chemical_bonds uid
                                subchild_lvl_8 = etree.SubElement(subchild_lvl_7, "uid")
                                fill_child_action(subchild_lvl_8, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{mode_number}_Bonds"][bonds_number])
                        # bands band assignments assignment transition vibration_modes comments
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                        fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Vibrations_Comment"])
                        # bands band assignments assignment transition ROTATION_comment
                        subchild_lvl_5 = etree.Comment(' TRANSITION: ROTATION MODES ')
                        subchild_lvl_4.insert(5, subchild_lvl_5)
                        # bands band assignments assignment transition rotation_modes
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "rotation_modes")
                        if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types"]) and is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Label"]):                            # bands band assignments assignment transition rotation_modes mode
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                            # bands band assignments assignment transition rotation_modes mode label
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                            fill_child_action(subchild_lvl_7, "")
                            # bands band assignments assignment transition rotation_modes mode type
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                            fill_child_action(subchild_lvl_7, "")
                        else:
                            for mode_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types"])):
                                if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types"][mode_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Label"][mode_number] != "":
                                    # bands band assignments assignment transition rotation_modes mode
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                                    # bands band assignments assignment transition rotation_modes mode label
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Label"][mode_number])
                                    # bands band assignments assignment transition rotation_modes mode type
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types"][mode_number])

                        # bands band assignments assignment transition rotation_modes comments
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                        fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Comment"])
                        # bands band assignments assignment transition PHONON_comment
                        subchild_lvl_5 = etree.Comment(' TRANSITION: PHONON MODES ')
                        subchild_lvl_4.insert(7, subchild_lvl_5)
                        # bands band assignments assignment transition phonon_modes
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "phonon_modes")
                        if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types"]) and is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Label"]):                            # bands band assignments assignment transition phonon_modes mode
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                            # bands band assignments assignment transition phonon_modes mode label
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                            fill_child_action(subchild_lvl_7, "")
                            # bands band assignments assignment transition phonon_modes mode type
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                            fill_child_action(subchild_lvl_7, "")
                        else:
                            for mode_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types"])):
                                if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types"][mode_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Label"][mode_number] != "":
                                    # bands band assignments assignment transition phonon_modes mode
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "mode")
                                    # bands band assignments assignment transition phonon_modes mode label
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "label")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Label"][mode_number])
                                    # bands band assignments assignment transition phonon_modes mode type
                                    subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                                    fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types"][mode_number])
                        # bands band assignments assignment transition phonon_modes comments
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                        fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Comment"])
                        # bands band assignments assignment RESONANCES_comment
                        subchild_lvl_4 = etree.Comment(' TRANSITION: RESONANCES ')
                        subchild_lvl_3.insert(13, subchild_lvl_4)
                        # bands band assignments assignment resonances
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "resonances")
                        resonances_qty = 0
                        for item_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"])):
                            if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"][item_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb"][item_number] != "":
                                resonances_qty = resonances_qty + 1
                        if resonances_qty == 0:
                            # bands band assignments assignment resonances resonance
                            subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "resonance")
                            # bands band assignments assignment resonances type
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "type")
                            fill_child_action(subchild_lvl_6, "")
                            # bands band assignments assignment resonances band_uid
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "band_uid")
                            fill_child_action(subchild_lvl_6, "")
                            # bands band assignments assignment resonances band_Assignment_number
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "band_assignment_number")
                            fill_child_action(subchild_lvl_6, "")
                            # bands band assignments assignment resonances comments
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                            fill_child_action(subchild_lvl_6, "")
                        else:
                            for item_number in range(0, len(XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"])):
                                if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"][item_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number] != "" or XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb"][item_number] != "":
                                    # bands band assignments assignment resonances resonance
                                    subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "resonance")
                                    # bands band assignments assignment resonances type
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "type")
                                    fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"][item_number])
                                    # bands band assignments assignment resonances band_uid
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "band_uid")
                                    if XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number] and XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number] != "NULL":
                                        fill_child_action(subchild_lvl_6, 'BAND_' + XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number])
                                    elif XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][item_number] == "NULL":
                                        fill_child_action(subchild_lvl_6, 'NULL')
                                    else:
                                        fill_child_action(subchild_lvl_6, '')
                                    # bands band assignments assignment resonances band_Assignment_number
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "band_assignment_number")
                                    fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb"][item_number])
                                    # bands band assignments assignment resonances comments
                                    subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "comments")
                                    fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Comment"][item_number])
                    # bands band CHARACTERISTICS_comment
                    subchild_lvl_2 = etree.Comment(' ENVIRONMENT AND CHARACTERISTICS ')
                    subchild_lvl_1.insert(5, subchild_lvl_2)
                    # bands band characteristics
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "characteristics")
                    for char_number in range(1, XLSX_data[f"B_{band_number}_Characteristics_qty"] + 1):
                        # bands band characteristics characteristic
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "characteristic")
                        # bands band characteristics characteristic number
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "number")
                        fill_child_action(subchild_lvl_4, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Nb"])
                        # bands band characteristics characteristic CONSTITUENT_comment
                        subchild_lvl_4 = etree.Comment(' CONSTITUENT ')
                        subchild_lvl_3.insert(1, subchild_lvl_4)
                        # bands band characteristics characteristic constituent
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "constituent")
                        # bands band characteristics characteristic constituent composition_comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "composition_comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Composition"])
                        # bands band characteristics characteristic constituent texture_comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "texture_comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Texture"])
                        # bands band characteristics characteristic constituent parameters_environment
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "parameters_environment")
                        # bands band characteristics characteristic constituent parameters_environment temperature
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "temperature")
                        # bands band characteristics characteristic constituent parameters_environment temperature unit
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "unit")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Unit"])
                        # bands band characteristics characteristic constituent parameters_environment temperature value
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "value")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Value"])
                        # bands band characteristics characteristic constituent parameters_environment temperature error
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "error")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Error"])
                        # bands band characteristics characteristic constituent parameters_environment temperature formation
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "formation")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Formation"])
                        # bands band characteristics characteristic constituent parameters_environment temperature max
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "max")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Max"])
                        # bands band characteristics characteristic constituent parameters_environment temperature comments
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "comments")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_T_Comment"])
                        # bands band characteristics characteristic constituent parameters_environment pressure
                        subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "pressure")
                        # bands band characteristics characteristic constituent parameters_environment pressure unit
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "unit")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Unit"])
                        # bands band characteristics characteristic constituent parameters_environment pressure value
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "value")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Value"])
                        # bands band characteristics characteristic constituent parameters_environment pressure error
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "error")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Error"])
                        # bands band characteristics characteristic constituent parameters_environment pressure formation
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "formation")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Formation"])
                        # bands band characteristics characteristic constituent parameters_environment pressure max
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "max")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Max"])
                        # bands band characteristics characteristic constituent parameters_environment pressure stress_type
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "stress_type")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Stress_type"])
                        # bands band characteristics characteristic constituent parameters_environment pressure comments
                        subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "comments")
                        fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_P_Comment"])
                        # bands band characteristics characteristic EXCITATION_comment
                        subchild_lvl_4 = etree.Comment(' EXCITATION LIGHT ')
                        subchild_lvl_3.insert(3, subchild_lvl_4)
                        # bands band characteristics characteristic excitation
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "excitation")
                        # bands band characteristics characteristic excitation laser_wavelength
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "laser_wavelength")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Laser_excitation_Wavelength"])
                        # bands band characteristics characteristic excitation sample_orientation_mode
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "sample_orientation_mode")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Sample_Orient_mode"])
                        # bands band characteristics characteristic excitation sample_orientation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "sample_orientation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Sample_Orient"])
                        # bands band characteristics characteristic excitation polarization_orientation_mode
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "polarization_orientation_mode")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Polarization_Orient_mode"])
                        # bands band characteristics characteristic excitation polarization_orientation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "polarization_orientation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Polarization_Orient"])
                        # bands band characteristics characteristic excitation comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Excitation_Comment"])
                        # bands band characteristics characteristic METHODS_comment
                        subchild_lvl_4 = etree.Comment(' CHARACTERISTICS METHODS ')
                        subchild_lvl_3.insert(5, subchild_lvl_4)
                        # bands band characteristics characteristic methods
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "methods")
                        for method_number in range(1, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Methods_qty"] + 1):
                            # bands band characteristics characteristic methods method
                            subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "method")
                            # bands band characteristics characteristic methods method type
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "type")
                            fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Method_{method_number}_Types"])
                            # bands band characteristics characteristic methods method description
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "description")
                            fill_child_action(subchild_lvl_6, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Method_{method_number}_Description"])
                            # bands band characteristics characteristic methods method fit_function
                            subchild_lvl_6 = etree.SubElement(subchild_lvl_5, "fit_function")
                            # bands band characteristics characteristic methods method fit_function type
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "type")
                            fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Method_{method_number}_Fit_Fct_type"])
                            # bands band characteristics characteristic methods method fit_function parameters
                            subchild_lvl_7 = etree.SubElement(subchild_lvl_6, "parameters")
                            fill_child_action(subchild_lvl_7, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Method_{method_number}_Fit_parameters"])
                        # bands band characteristics characteristic OVERLAP_comment
                        subchild_lvl_4 = etree.Comment(' OVERLAP ')
                        subchild_lvl_3.insert(7, subchild_lvl_4)
                        # bands band characteristics characteristic overlap
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "overlap")
                        fill_child_action(subchild_lvl_4, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Methods_Overlap"])
                        # bands band characteristics characteristic POSITION_comment
                        subchild_lvl_4 = etree.Comment(' POSITION ')
                        subchild_lvl_3.insert(9, subchild_lvl_4)
                        # bands band characteristics characteristic position
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "position")
                        # bands band characteristics characteristic position peak_method
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "peak_method")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Peak_method"])
                        # bands band characteristics characteristic position peak
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "peak")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Peak"])
                        # bands band characteristics characteristic position peak_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "peak_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Peak_error"])
                        # bands band characteristics characteristic position center_method
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "center_method")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Center_method"])
                        # bands band characteristics characteristic position center
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "center")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Center"])
                        # bands band characteristics characteristic position center_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "center_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Center_error"])
                        # bands band characteristics characteristic position evaluation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "evaluation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Evaluation"])
                        # bands band characteristics characteristic position comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Position_Comment"])
                        # bands band characteristics characteristic WIDTH_comment
                        subchild_lvl_4 = etree.Comment(' WIDTH ')
                        subchild_lvl_3.insert(11, subchild_lvl_4)
                        # bands band characteristics characteristic width
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "width")
                        # bands band characteristics characteristic width method
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "method")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Method"])
                        # bands band characteristics characteristic width fwhm
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "fwhm")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_FWHM"])
                        # bands band characteristics characteristic width fwhm_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "fwhm_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_FWHM_error"])
                        # bands band characteristics characteristic width shape
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "shape")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Shape"])
                        # bands band characteristics characteristic width asymmetry_factor
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "asymmetry_factor")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Asymm_factor"])
                        # bands band characteristics characteristic width asymmetry_factor_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "asymmetry_factor_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Asymm_factor_error"])
                        # bands band characteristics characteristic width evaluation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "evaluation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Evaluation"])
                        # bands band characteristics characteristic width comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Width_Comments"])
                        # bands band characteristics characteristic INTENSITY_comment
                        subchild_lvl_4 = etree.Comment(' INTENSITY ')
                        subchild_lvl_3.insert(13, subchild_lvl_4)
                        # bands band characteristics characteristic peak_intensity
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "peak_intensity")
                        # bands band characteristics characteristic peak_intensity method
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "method")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Method"])
                        # bands band characteristics characteristic peak_intensity abscoef
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"])
                        # bands band characteristics characteristic peak_intensity abscoef_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef_error"])
                        # bands band characteristics characteristic peak_intensity abscoef_specific
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_specific")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef_sp"])
                        # bands band characteristics characteristic peak_intensity abscoef_specific_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_specific_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef_sp_error"])
                        # bands band characteristics characteristic peak_intensity relative
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "relative")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"])
                        # bands band characteristics characteristic peak_intensity relative_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "relative_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative_error"])
                        # bands band characteristics characteristic peak_intensity strength
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "strength")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Strength"])
                        # bands band characteristics characteristic peak_intensity evaluation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "evaluation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Evaluation"])
                        # bands band characteristics characteristic peak_intensity comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Comment"])
                        # bands band characteristics characteristic INTEGRATED_INTENSITY_comment
                        subchild_lvl_4 = etree.Comment(' INTEGRATED INTENSITY ')
                        subchild_lvl_3.insert(15, subchild_lvl_4)
                        # bands band characteristics characteristic integrated_intensity
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "integrated_intensity")
                        # bands band characteristics characteristic integrated_intensity method
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "method")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Method"])
                        # bands band characteristics characteristic integrated_intensity abscoef
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"])
                        # bands band characteristics characteristic integrated_intensity abscoef_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef_error"])
                        # bands band characteristics characteristic integrated_intensity abscoef_specific
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_specific")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef_sp"])
                        # bands band characteristics characteristic integrated_intensity abscoef_specific_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "abscoef_specific_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef_sp_error"])
                        # bands band characteristics characteristic integrated_intensity relative
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "relative")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"])
                        # bands band characteristics characteristic integrated_intensity relative_error
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "relative_error")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative_error"])
                        # bands band characteristics characteristic integrated_intensity strength
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "strength")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Strength"])
                        # bands band characteristics characteristic integrated_intensity evaluation
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "evaluation")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Evaluation"])
                        # bands band characteristics characteristic integrated_intensity comments
                        subchild_lvl_5 = etree.SubElement(subchild_lvl_4, "comments")
                        fill_child_action(subchild_lvl_5, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Comment"])
                        # bands band characteristics characteristic PREVIEW_comment
                        subchild_lvl_4 = etree.Comment(' BANDLIST PREVIEW ')
                        subchild_lvl_3.insert(17, subchild_lvl_4)
                        # bands band characteristics characteristic bandlist_nominal_flag
                        subchild_lvl_4 = etree.SubElement(subchild_lvl_3, "bandlist_nominal_flag")
                        fill_child_action(subchild_lvl_4, XLSX_data[f"B_{band_number}_Characteristic_{char_number}_Bandlist_flag"])
                    # bands band REFERENCES_comment
                    subchild_lvl_2 = etree.Comment(' REFERENCES ')
                    subchild_lvl_1.insert(7, subchild_lvl_2)
                    # bands band publications
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "publications")
                    if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Publications_UID"]):
                        # bands band publications publication_uid
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "publication_uid")
                        fill_child_action(subchild_lvl_3, "NULL")
                    else:
                        for item in XLSX_data[f"B_{band_number}_Publications_UID"]:
                            # bands band publications publication_uid
                            subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "publication_uid")
                            if item == "NULL" or item.find('PUBLI_') != -1:
                                fill_child_action(subchild_lvl_3, item)
                            else:
                                fill_child_action(subchild_lvl_3, 'PUBLI_' + item)
                    # bands band data_publication_spectra
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "data_publication_spectra")
                    if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Publications_SSHADE_UID"]):
                        # bands band data_publication_spectra experiment_uid
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "spectrum_uid")
                        fill_child_action(subchild_lvl_3, "")
                    else:
                        for item in XLSX_data[f"B_{band_number}_Publications_SSHADE_UID"]:
                            # bands band data_publication_experiments experiment_uid
                            subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "spectrum_uid")
                            if item.find('SPECTRUM_') != -1:
                                fill_child_action(subchild_lvl_3, item)
                            else:
                                fill_child_action(subchild_lvl_3, 'SPECTRUM_' + item)
                    # bands band data_publication_links
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "data_publication_links")
                    if is_list_with_only_empty_strings(XLSX_data[f"B_{band_number}_Publications_Data_URL"]):
                        # bands band data_publication_experiments link
                        subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "link")
                        fill_child_action(subchild_lvl_3, "")
                    else:
                        for item in XLSX_data[f"B_{band_number}_Publications_Data_URL"]:
                            # bands band data_publication_experiments experiment_uid
                            subchild_lvl_3 = etree.SubElement(subchild_lvl_2, "link")
                            fill_child_action(subchild_lvl_3, item)
                    # bands band publication_comments
                    subchild_lvl_2 = etree.SubElement(subchild_lvl_1, "publication_comments")
                    fill_child_action(subchild_lvl_2, XLSX_data[f"B_{band_number}_Publications_Comments"])
    # from xml to byte
    str_to_upload = etree.tostring(xml_root, pretty_print=True, encoding="utf-8", xml_declaration=True, method="xml")
    return str_to_upload, uid


# function to verify abs_mandatory and mandatory
def verification_F(xlsx_workbook, bandlist_type):
    # inner function to verify
    def verif_action(data_to_verif: dict, data_position: dict, key_word: str, tag_names_list: list, sheet_name: str, mand_flag: bool, index=-1) -> str:
        """
        This function takes data from an xlsx_workbook, verifies it
        and returns an empty string "" if the data is Ok
        or a string with part of warning message in the opposite case.

        This function also fills Mandatory fields in the data_to_verif with "NULL" if they have no values.

        This function takes data_to_verif for dict with the result of XLSX reading,
        data_position for dict with positions for data in the XLSX file,
        key_word for the key in the data_to_verif and data_position dicts,
        tag_names_list to show it in the warning message,
        sheet_name for "BandList" or "Band" to show it in the warning message,
        mand_flag is False for ABS_mandatory or True for Mandatory,
        and index for cases when value in the data_to_verif is an array.
        """
        # determines the position to show
        position = ""
        if data_position[key_word][1] == data_position[key_word][2]:
            position = data_position[key_word][0] + str(data_position[key_word][1])
        else:
            position = data_position[key_word][0] + str(data_position[key_word][1]) + "-" + str(data_position[key_word][2])
        # determines the tag to show
        tag_name = ""
        last_tag = tag_names_list[-1]
        for element in tag_names_list:
            tag_name = tag_name + "<" + element + ">"
        # verification and filling with "NULL" for Mandatory
        if not data_to_verif[key_word]:
            if mand_flag:
                data_to_verif[key_word] = "NULL"
            if position:
                return f"- {tag_name}: {sheet_name}, {position}\n"
            else:
                return f"- {tag_name}: {sheet_name}: no {last_tag}\n"
        else:
            if type(data_to_verif[key_word]) is list:
                if data_to_verif[key_word] == [""] or is_list_with_only_empty_strings(data_to_verif[key_word]):
                    if mand_flag:
                        data_to_verif[key_word] = ["NULL"]
                    if position:
                        return f"- {tag_name}: {sheet_name}, {position}\n"
                    else:
                        return f"- {tag_name}: {sheet_name}: no {last_tag}\n"
                else:
                    if index != -1:
                        if data_to_verif[key_word][index] == "":
                            if mand_flag:
                                data_to_verif[key_word][index] = "NULL"
                            position = data_position[key_word][0] + str(data_position[key_word][1] + index)
                            if position:
                                return f"- {tag_name}: {sheet_name}, {position}\n"
                            else:
                                return f"- {tag_name}: {sheet_name}: no {last_tag}\n"
                        else:
                            return ""
                    else:
                        return ""
            else:
                if not mand_flag and data_to_verif[key_word] == "NULL" and position:
                    return f"- {tag_name}: {sheet_name}, NULL in ABS_M in {position}\n"
                else:
                    return ""

    str_list_problems = ""
    data_read = XLSX_reader(xlsx_workbook, bandlist_type)
    data_to_verif = data_read[0]
    data_position = data_read[1]
    no_assignments = False
    no_chars = False
    # ABS_mandatory
    abs_mand_problems = "ABS_MANDATORY list:\n"
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Import_mode", ["import_mode"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Type", ["type"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Title", ["title"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Export_filename", ["export_filename"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_UID", ["uid"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Constituent_UID", ["constituent", "uid"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Constituent_Primary_specie_UID", ["constituent", "primary_specie_uid"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Unit", ["parameters_spectral", "unit"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Standard", ["parameters_spectral", "standard"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Range_types", ["parameters_spectral", "range_types", "type"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Range_min", ["parameters_spectral", "ranges", "min"], "BandList", False)
    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Range_max", ["parameters_spectral", "ranges", "max"], "BandList", False)
    if data_to_verif["BL_Import_mode"] == 'new version' or data_to_verif["BL_Import_mode"] == 'invalidate':
        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Versions_Previous_version_status", ["previous_version", "status"], "BandList", False)
        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, "BL_Versions_Comments", ["previous_version", "comments"], "BandList", False)
    for section_number in range(1, data_to_verif["BL_Sections_qty"] + 1):
        if data_to_verif[f"BL_Section_{section_number}_Bands_UID"] == [""] and data_to_verif[f"BL_Section_{section_number}_Sub_section_1_Bands_UID"] == [""]:
            abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"BL_Section_{section_number}_Bands_UID", ["structure", "sections", "band_uid"], f"BandList, section_{section_number}", False)
    for band_number in range(1, data_to_verif["B_qty"] + 1):
        current_band_number = data_to_verif[f'B_{band_number}_Index']
        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Import_mode", ["bands", "import_mode"], f"Band_{current_band_number}", False)  # never used
        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_UID", ["bands", "uid"], f"Band_{current_band_number}", False)
        no_assignments = False
        if data_to_verif[f"B_{band_number}_Assignments_qty"] == 0 or (data_to_verif[f"B_{band_number}_Assignments_qty"] == 1 and f"B_{band_number}_Assignment_1_Number" in data_to_verif.keys() and data_to_verif[f"B_{band_number}_Assignment_1_Number"] == ""):
            no_assignments = True
            abs_mand_problems = abs_mand_problems + f"- <bands><assignments>: Band_{current_band_number}, No assignment found in {data_position[f'B_{band_number}_Assignment_1_Number'][0]}{data_position[f'B_{band_number}_Assignment_1_Number'][1]}\n"
        if not no_assignments:
            for assignment_number in range(1, data_to_verif[f"B_{band_number}_Assignments_qty"] + 1):
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Label", ["bands", "assignments", "label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", False)
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Category", ["bands", "assignments", "category"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", False)
                for multy_index in range(0, len(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"])):
                    if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Degeneracy"][multy_index] != "" or data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band"][multy_index] != "":
                        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types", ["bands", "assignments", "multiplicity", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", False, multy_index)
                for specie_number in range(1, data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Transition_Species_qty"] + 1):
                    if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] != "phonon mode":
                        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Transition_Specie_{specie_number}_UID", ["bands", "assignments", "primary_specie", "uid"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", False)
                if (is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types"]) and is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Electronic_Labels"])) and (is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types"]) and is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Rotation_Label"])) and (is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types"]) and is_list_with_only_empty_strings(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Phonon_Label"])):
                    vib_flag = 0
                    for vib_number in range(1, data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Vibrations_qty"] + 1):
                        if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{vib_number}_Types"] == "" and data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Vibration_{vib_number}_Label"] == "":
                            vib_flag = vib_flag + 1
                    if vib_flag == data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Vibrations_qty"]:
                        abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Vibration_1_Types", ["bands", "assignments", "transition"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", False)
        no_chars = False
        if data_to_verif[f"B_{band_number}_Characteristics_qty"] == 0 or (data_to_verif[f"B_{band_number}_Characteristics_qty"] == 1 and f"B_{band_number}_Characteristic_1_Nb" in data_to_verif.keys() and data_to_verif[f"B_{band_number}_Characteristic_1_Nb"] == ""):
            no_chars = True
            abs_mand_problems = abs_mand_problems + f"- <bands><characteristics>: Band_{band_number}, No characteristic found\n"
        if not no_chars:
            for char_number in range(1, data_to_verif[f"B_{band_number}_Characteristics_qty"] + 1):
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_T_Unit", ["bands", "characteristics", "temperature", "unit"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_T_Value", ["bands", "characteristics", "temperature", "value"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Value"] != "" or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Formation"] != "" or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Max"] != "":
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_P_Unit", ["bands", "characteristics", "pressure", "unit"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Position_Peak"]:
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Peak_method", ["bands", "characteristics", "position", "peak_method"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Peak_error", ["bands", "characteristics", "position", "peak_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                if not data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Position_Center"]:
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Peak", ["bands", "characteristics", "position", "peak"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                if not data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Position_Peak"]:
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Center", ["bands", "characteristics", "position", "center"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Position_Center"]:
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Center_method", ["bands", "characteristics", "position", "center_method"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                    abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Center_error", ["bands", "characteristics", "position", "center_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Position_Evaluation", ["bands", "characteristics", "position", "evaluation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
                abs_mand_problems = abs_mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Bandlist_flag", ["bands", "characteristics", "bandlist_nominal_flag"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", False)
    if abs_mand_problems != "ABS_MANDATORY list:\n":
        str_list_problems = str_list_problems + abs_mand_problems
    # Mandatory
    electronic_flag = 0
    infrared_flag = 0
    ref_position_IR_flag = 0
    mand_problems = "MANDATORY list:\n"
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Description", ["description"], "BandList", True)
    if data_to_verif["BL_Type"] in ["absorption", "Raman"]:
        infrared_flag = 1
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Analysis", ["analysis"], "BandList", True)
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Validation_Quality", ["quality_flag"], "BandList", True)
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Validation_Date_validated", ["date_validated"], "BandList", True)
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Validation_Validators_UID", ["validators"], "BandList", True)
    mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Sections_Var_param", ["structure", "sections variable_parameter"], "BandList, sections", True)
    for section_number in range(1, data_to_verif["BL_Sections_qty"] + 1):
        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"BL_Section_{section_number}_Title", ["structure", "sections", "title"], f"BandList, section_{section_number}", True)
        for sub_section_number in range(1, data_to_verif[f"BL_Section_{section_number}_Sub_sections_qty"] + 1):
            if not is_list_with_only_empty_strings(data_to_verif[f"BL_Section_{section_number}_Sub_section_{sub_section_number}_Bands_UID"]):
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"BL_Section_{section_number}_Sub_section_{sub_section_number}_Title", ["structure", "subsections", "title"], f"BandList, section_{section_number}, sub_section_{sub_section_number}", True)
    for band_number in range(1, data_to_verif["B_qty"] + 1):
        current_band_number = data_to_verif[f'B_{band_number}_Index']
        if data_to_verif[f"B_{band_number}_Publications_UID"] != ["NULL"] and data_to_verif[f"B_{band_number}_Publications_SSHADE_UID"] in [[''], ['NULL']] and data_to_verif[f"B_{band_number}_Publications_Data_URL"] in [[''], ['NULL']]:
            mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Publications_UID", ["bands", "publications"], f"Band_{current_band_number}", True)
        no_assignments = False
        if data_to_verif[f"B_{band_number}_Assignments_qty"] == 0 or (data_to_verif[f"B_{band_number}_Assignments_qty"] == 1 and f"B_{band_number}_Assignment_1_Number" in data_to_verif.keys() and data_to_verif[f"B_{band_number}_Assignment_1_Number"] == ""):
            no_assignments = True
        if not no_assignments:
            for assignment_number in range(1, data_to_verif[f"B_{band_number}_Assignments_qty"] + 1):
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Level", ["bands", "assignments", "level"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] in ["fundamental vibration", "rotation", "phonon mode"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Symmetry", ["bands", "assignments", "symmetry_label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Evaluation", ["bands", "assignments", "evaluation"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                contribution_flag = 0
                for multy_index in range(0, len(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"])):
                    if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"][multy_index] or data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Degeneracy"][multy_index] or data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band"][multy_index]:
                        if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Contribution_Level"] == "extracted":
                            mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Other_band", ["bands", "assignments", "multiplicity", "other_band_uid"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True, multy_index)
                        if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"][multy_index] in ["mode degeneracy", "site degeneracy", "accidental degeneracy"]:
                            mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Degeneracy", ["bands", "assignments", "multiplicity", "degeneracy"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True, multy_index)
                        if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"][multy_index] not in ["no", "mode", "site degeneracy", "unknown", ""] and data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Multiplicity_Types"][multy_index] in ["accidental degeneracy", "other isotope of primary specie", "other constituent specie", "other", ""]:
                            contribution_flag = 1
                if contribution_flag:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Contribution_Level", ["bands", "assignments", "contribution_level"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] == "electronic transition":
                    electronic_flag = 1
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Electronic_Labels", ["bands", "assignments", "transition", "electronic_modes", "label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Electronic_Types", ["bands", "assignments", "transition", "electronic_modes", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] in ["fundamental vibration", "overtone vibration", "combination vibration"]:
                    for vib_number in range(1, data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Vibrations_qty"] + 1):
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Vibration_{vib_number}_Label", ["bands", "assignments", "transition", "vibration_modes", "label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Vibration_{vib_number}_Types", ["bands", "assignments", "transition", "vibration_modes", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Vibration_{vib_number}_Bonds", ["bands", "assignments", "transition", "vibration_modes", "bonds"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] in ["rotation", "overtone rotation"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Rotation_Label", ["bands", "assignments", "transition", "rotation_modes", "label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Rotation_Types", ["bands", "assignments", "transition", "rotation_modes", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Category"] == "phonon mode":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Phonon_Label", ["bands", "assignments", "transition", "phonon_modes", "label"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Phonon_Types", ["bands", "assignments", "transition", "phonon_modes", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True)
                for resonance_number in range(0, len(data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"])):
                    if data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types"][resonance_number] or (data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][resonance_number] and data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band"][resonance_number] != "NULL") or (data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb"][resonance_number] and data_to_verif[f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb"][resonance_number] != "NULL"):
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Resonances_Types", ["bands", "assignments", "resonances", "type"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True, resonance_number)
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Resonances_Band", ["bands", "assignments", "resonances", "band_uid"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True, resonance_number)
                        mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Assignment_{assignment_number}_Resonances_Nb", ["bands", "assignments", "resonances", "band_assignment_number"], f"Band_{current_band_number}, assignment_{data_to_verif[f'B_{band_number}_Assignment_{assignment_number}_Number']}", True, resonance_number)
        no_chars = False
        if data_to_verif[f"B_{band_number}_Characteristics_qty"] == 0 or (data_to_verif[f"B_{band_number}_Characteristics_qty"] == 1 and f"B_{band_number}_Characteristic_1_Nb" in data_to_verif.keys() and data_to_verif[f"B_{band_number}_Characteristic_1_Nb"] == ""):
            no_chars = True
        if not no_chars:
            for char_number in range(1, data_to_verif[f"B_{band_number}_Characteristics_qty"] + 1):
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"]:
                    if electronic_flag:
                        electronic_flag = 2
                    if infrared_flag:
                        infrared_flag = 2
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_T_Error", ["bands", "characteristics", "temperature", "error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_T_Max", ["bands", "characteristics", "temperature", "max"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Error"] or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Formation"] or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Max"] or data_to_verif[f"B_{band_number}_Characteristic_{char_number}_P_Stress_type"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_P_Value", ["bands", "characteristics", "pressure", "value"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] in ["Raman scattering", "fluorescence emission"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Laser_excitation_Wavelength", ["bands", "characteristics", "excitation", "laser_wavelength"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] in ["Raman scattering"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Sample_Orient_mode", ["bands", "characteristics", "excitation", "sample_orientation_mode"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] in ["Raman scattering"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Sample_Orient_mode"] == "oriented":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Sample_Orient", ["bands", "characteristics", "excitation", "sample_orientation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] in ["Raman scattering", "absorption"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Sample_Orient_mode"] == "oriented":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Polarization_Orient_mode", ["bands", "characteristics", "excitation", "polarization_orientation_mode"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Polarization_Orient_mode"] == "polarized":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Polarization_Orient", ["bands", "characteristics", "excitation", "polarization_orientation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Method_1_Types", ["bands", "characteristics", "method", "type"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Methods_Overlap", ["bands", "characteristics", "overlap"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_FWHM", ["bands", "characteristics", "width", "fwhm"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Width_FWHM"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Width_FWHM"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_Method", ["bands", "characteristics", "width", "method"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_FWHM_error", ["bands", "characteristics", "width", "fwhm_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_Evaluation", ["bands", "characteristics", "width", "evaluation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_Shape", ["bands", "characteristics", "width", "shape"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Width_Shape"] in ["asymmetric", "asymmetric low frequency wing", "asymmetric high frequency wing"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Width_Asymm_factor", ["bands", "characteristics", "width", "asymmetry_factor"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Strength"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Strength"] != "NULL"):
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Method", ["bands", "characteristics", "peak_intensity", "method"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] == "absorption" and not data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef", ["bands", "characteristics", "peak_intensity", "abscoef"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef_error", ["bands", "characteristics", "peak_intensity", "abscoef_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] == [""] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] != [""]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative", ["bands", "characteristics", "peak_intensity", "relative"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] != [""] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] == [""]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef", ["bands", "characteristics", "peak_intensity", "abscoef"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative_error", ["bands", "characteristics", "peak_intensity", "relative_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Strength", ["bands", "characteristics", "peak_intensity", "strength"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Abs_coef"] != "NULL"):
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Evaluation", ["bands", "characteristics", "peak_intensity", "evaluation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Strength"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Strength"] != "NULL"):
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Method", ["bands", "characteristics", "integrated_intensity", "method"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif["BL_Type"] == "absorption" and not data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"]:
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef", ["bands", "characteristics", "integrated_intensity", "abscoef"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef_error", ["bands", "characteristics", "integrated_intensity", "abscoef_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                elif data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative", ["bands", "characteristics", "integrated_intensity", "relative"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] != "NULL":
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative_error", ["bands", "characteristics", "integrated_intensity", "relative_error"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Abs_coef"] != "NULL"):
                    mand_problems = mand_problems + verif_action(data_to_verif, data_position, f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Evaluation", ["bands", "characteristics", "integrated_intensity", "evaluation"], f"Band_{current_band_number}, characteristic_{data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Nb']}", True)
                if not ref_position_IR_flag and ((data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Peak_intensity_Relative"] != "NULL") or (data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] and data_to_verif[f"B_{band_number}_Characteristic_{char_number}_Integrated_intensity_Relative"] != "NULL")):
                    ref_position_IR_flag = 1
    if electronic_flag == 2:
        mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Ref_pos_electronic", ["reference_position", "electronic"], "BandList", True)
    if infrared_flag == 2 or ref_position_IR_flag:
        mand_problems = mand_problems + verif_action(data_to_verif, data_position, "BL_Spectral_Ref_pos_absorption", ["reference_position", "infrared"], "BandList", True)
    if mand_problems != "MANDATORY list:\n":
        if str_list_problems:
            str_list_problems = str_list_problems + "\n"
        str_list_problems = str_list_problems + mand_problems
    # Oubliettes
    oubl_problems = "OUBLIETTES list:\n"
    if xlsx_workbook[xlsx_workbook.rfind("/") + 1:] != data_to_verif["BL_Original_data_filename"]:
        oubl_problems = oubl_problems + "- <original_data_filename>: C12 value doesn't correspond to this xlsx file name\n"
    for section_number in range(1, data_to_verif["BL_Sections_qty"] + 1):
        # at least one section title
        if f"BL_Section_{section_number}_Sub_section_1_title" in data_to_verif.keys() and data_to_verif[f"BL_Section_{section_number}_Sub_section_1_title"] != "":
            oubl_problems = oubl_problems + verif_action(data_to_verif, data_position, f"BL_Section_{section_number}_Var_param", ["structure", "sections", "var_param"], f"BandList, section_{section_number}", False)
        # structure bands UID in bans UID
        # for section UID
        for band_UID_item in data_to_verif[f"BL_Section_{section_number}_Bands_UID"]:
            if band_UID_item:
                band_UID_flag = 0
                for band_number in range(1, data_to_verif["B_qty"] + 1):
                    if band_UID_item == data_to_verif[f"B_{band_number}_UID"]:
                        band_UID_flag = band_UID_flag + 1
                if band_UID_flag == 0:
                    oubl_problems = oubl_problems + f"- <structure><section><bands>: Section_{section_number}, no band UID for Section Band UID {band_UID_item}\n"
                elif band_UID_flag > 1:
                    oubl_problems = oubl_problems + f"- <structure><section><bands>: Section_{section_number}, more than one band UID for Section Band UID {band_UID_item}\n"
        # for sub_section UID
        for sub_section_number in range(1, data_to_verif[f"BL_Section_{section_number}_Sub_sections_qty"] + 1):
            for band_UID_item in data_to_verif[f"BL_Section_{section_number}_Sub_section_{sub_section_number}_Bands_UID"]:
                if band_UID_item:
                    band_UID_flag = 0
                    for band_number in range(1, data_to_verif["B_qty"] + 1):
                            if band_UID_item == data_to_verif[f"B_{band_number}_UID"]:
                                band_UID_flag = band_UID_flag + 1
                    if band_UID_flag == 0:
                        oubl_problems = oubl_problems + f"- <structure><section><bands>: Section_{section_number}, sub_section_{sub_section_number}, no band UID for Section Band UID {band_UID_item}\n"
                    elif band_UID_flag > 1:
                        oubl_problems = oubl_problems + f"- <structure><section><bands>: Section_{section_number}, sub_section_{sub_section_number}, more than one band UID for Section Band UID {band_UID_item}\n"
    for band_number in range(1, data_to_verif["B_qty"] + 1):
        current_band_number = data_to_verif[f'B_{band_number}_Index']
        # nominal flag
        nominal_flags = 0
        for char_number in range(1, data_to_verif[f'B_{band_number}_Characteristics_qty'] + 1):
            # no <laser_wavelength> for bandlist_type = 'absorption'
            if data_to_verif[f'BL_Type'] == 'absorption' and data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Laser_excitation_Wavelength']:
                oubl_problems = oubl_problems + f"- <bands><characteristics><excitation><laser_wavelength>: Band_{current_band_number}, non-empty value when bandlist_type is 'absorption'\n"
            # nominal flag
            if data_to_verif[f'B_{band_number}_Characteristic_{char_number}_Bandlist_flag'].lower() in ['yes', 'true']:
                nominal_flags = nominal_flags + 1
        if nominal_flags > 1:
            oubl_problems = oubl_problems + f"- <bands><characteristics><nominal_flag>: Band_{current_band_number}, More than one nominal_flag\n"
        # bands UID in structure
        band_UID_flag = 0
        for section_number in range(1, data_to_verif["BL_Sections_qty"] + 1):
            if data_to_verif[f"B_{band_number}_UID"] in data_to_verif[f"BL_Section_{section_number}_Bands_UID"]:
                band_UID_flag = band_UID_flag + 1
            for sub_section_number in range(1, data_to_verif[f"BL_Section_{section_number}_Sub_sections_qty"] + 1):
                for sub_band_uid in data_to_verif[f"BL_Section_{section_number}_Sub_section_{sub_section_number}_Bands_UID"]:
                    if data_to_verif[f"B_{band_number}_UID"] == sub_band_uid:
                        band_UID_flag = band_UID_flag + 1
        if band_UID_flag == 0:
            oubl_problems = oubl_problems + f"- <bands><band><uid>: Band_{current_band_number}, no Band UID in BandList structure\n"
        elif band_UID_flag > 1:
            oubl_problems = oubl_problems + f"- <bands><band><uid>: Band_{current_band_number}, more than one Band UID in BandList structure\n"
    # bands UID in structure: from bandlist
    for section_number in range(1, data_to_verif["BL_Sections_qty"] + 1):
        if len(data_to_verif[f"BL_Section_{section_number}_Bands_UID"]) != len(set(data_to_verif[f"BL_Section_{section_number}_Bands_UID"])):
            oubl_problems = oubl_problems + f"- <structure><bands>: Section_{section_number}, more than one Band UID in BandList structure\n"
        for sub_section_index in range(1, data_to_verif[f"BL_Section_{section_number}_Sub_sections_qty"] + 1):
            if len(data_to_verif[f"BL_Section_{section_number}_Sub_section_{sub_section_index}_Bands_UID"]) != len(set(data_to_verif[f"BL_Section_{section_number}_Sub_section_{sub_section_index}_Bands_UID"])):
                oubl_problems = oubl_problems + f"- <structure><subsections><bands>: Section_{section_number}, Sub-section_{sub_section_index}, more than one Band UID in BandList structure\n"
    if oubl_problems != "OUBLIETTES list:\n":
        if str_list_problems:
            str_list_problems = str_list_problems + "\n"
        str_list_problems = str_list_problems + oubl_problems
    return str_list_problems


# Demo data
def Demo_F():
    # INPUT
    # demo data
    #xlsx_workbook = "resources/exemples/bandlist_ABS_test_v092a.xlsx"
    #xlsx_workbook = "resources/exemples/bandlist_ABS_Raman_test_v092a.xlsx"
    #xlsx_workbook = "resources/exemples/bandlist_ABS_Raman_full_v092a.xlsx"
    xlsx_workbook = "tests/xlsx/verifications/mand/bandlist_section.xlsx"
    # bandlist type: ABS or RAMAN
    #bandlist_type = "RAMAN"
    bandlist_type = "RAMAN"

    # PROCESSING
    read_reult = XLSX_reader(xlsx_workbook, bandlist_type)
    #print("data_reading", read_reult[0])
    #print("position", read_reult[1])

    str_to_upload = XML_filler(xlsx_workbook, bandlist_type)[0]

    verification_result = verification_F(xlsx_workbook, bandlist_type)

    # OUTPUT
    print(verification_result)

    # xml file saving
    file_name = f"bandlist.xml"
    if file_name:
        with open(file_name, 'wb') as file_output:
            file_output.write(str_to_upload)


# run Demo data to test
#Demo_F()
